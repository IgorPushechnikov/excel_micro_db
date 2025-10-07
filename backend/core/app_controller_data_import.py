# backend/core/app_controller_data_import.py
"""
Модуль, содержащий функции для импорта различных типов данных из Excel-файла
в БД проекта через AppController.

Функции предназначены для вызова из AppController для реализации
импорта "по типам" (данные, стили, диаграммы, формулы) и "по режимам"
(всё, выборочно, частями).

ПРИМЕЧАНИЕ: Функции теперь принимают экземпляр ProjectDBStorage напрямую,
чтобы обеспечить потокобезопасность.
"""

import logging
import os
from typing import Dict, Any, List, Optional, Union
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
import numpy as np
# Исправлено: Импорт get_column_letter напрямую
from openpyxl.utils import get_column_letter

# Импортируем logger из utils
from backend.utils.logger import get_logger

# Импортируем функции из analyzer
# Исправлено: Импорт из правильного модуля и с правильными именами
from backend.analyzer.logic_documentation import _serialize_style, _serialize_chart

# Импортируем ProjectDBStorage
from backend.storage.base import ProjectDBStorage

logger = get_logger(__name__)


# --- Вспомогательные функции ---

# УДАЛЯЕМ _get_sheet_id_by_name, так как storage.save_sheet делает то же самое лучше.
# def _get_sheet_id_by_name(storage, sheet_name: str) -> Optional[int]:
#     """
#     Вспомогательная функция для получения sheet_id по имени листа.
#     """
#     if not storage or not storage.connection:
#         return None
#
#     try:
#         cursor = storage.connection.cursor()
#         # Предполагаем project_id = 1
#         cursor.execute("SELECT sheet_id FROM sheets WHERE name = ? AND project_id = 1", (sheet_name,))
#         result = cursor.fetchone()
#         return result[0] if result else None
#     except Exception as e:
#         logger.error(f"Ошибка при получении sheet_id для листа '{sheet_name}': {e}")
#         return None


# --- Функции для импорта "всё" по типам ---

def import_raw_data_from_excel(storage: ProjectDBStorage, file_path: str, options: Optional[Dict[str, Any]] = None) -> bool:
    """
    Импортирует только "сырые" данные (значения ячеек) из Excel-файла в БД проекта.
    Использует openpyxl с data_only=False для получения формул как строк.
    Импортирует частями.

    Args:
        storage: Экземпляр ProjectDBStorage для сохранения данных.
        file_path (str): Путь к Excel-файлу для импорта.
        options (Optional[Dict[str, Any]]): Опции импорта.
            {
                'chunk_size_rows': int, # Количество строк в одной части (по умолчанию 50)
                'sheets': List[str]     # Список имён листов для импорта. Если пуст, все.
            }

    Returns:
        bool: True, если импорт прошёл успешно, иначе False.
    """
    if not storage:
        logger.error("Экземпляр ProjectDBStorage не предоставлен. Невозможно выполнить импорт.")
        return False

    if not os.path.exists(file_path):
        logger.error(f"Excel-файл для импорта не найден: {file_path}")
        return False

    logger.info(f"Начало импорта 'сырых' данных из Excel-файла: {file_path}")

    try:
        # --- НОВОЕ: Обработка ошибки Nested.from_tree ---
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=False) # <-- data_only=False для формул
        except TypeError as e:
            if "Nested.from_tree() missing 1 required positional argument: 'node'" in str(e):
                logger.error(f"Ошибка openpyxl при открытии файла '{file_path}': {e}")
                logger.warning("Файл может содержать неподдерживаемые структуры (например, pivot-таблицы). Импорт прерван.")
                return False
            else:
                raise # Если это другая ошибка TypeError, пробрасываем её
        # --- КОНЕЦ НОВОГО ---
        logger.debug(f"Книга '{file_path}' успешно открыта.")

        sheets_to_import_orig = options.get('sheets', []) if options else []
        if not sheets_to_import_orig:
            sheets_to_import_orig = workbook.sheetnames

        # Явно приводим элементы к str для устранения ошибки Pylance
        sheets_to_import: List[str] = [str(name) for name in sheets_to_import_orig]

        for sheet_name_orig in sheets_to_import:
            # Убедимся, что имя листа - строка
            sheet_name: str = str(sheet_name_orig)
            if sheet_name not in workbook.sheetnames:
                logger.warning(f"Лист '{sheet_name}' не найден в файле '{file_path}'. Пропущен.")
                continue

            logger.info(f"Импорт 'сырых' данных с листа: {sheet_name}")
            
            # --- НОВОЕ: Гарантируем, что запись о листе существует ---
            # Предполагаем project_id = 1 для MVP
            sheet_id = storage.save_sheet(project_id=1, sheet_name=sheet_name)
            if sheet_id is None:
                logger.error(f"Не удалось создать/получить ID для листа '{sheet_name}'. Пропущен.")
                return False # Возвращаем False при ошибке создания записи о листе
            # --- КОНЕЦ НОВОГО ---

            sheet: Worksheet = workbook[sheet_name]
            total_rows = sheet.max_row
            chunk_size = options.get('chunk_size_rows', 50) if options else 50 # <-- ИЗМЕНЕНО С 100 НА 50
            
            # Исправление: Обработка случая, когда лист пуст (max_row возвращает None)
            if total_rows is None:
                logger.info(f"Лист '{sheet_name}' пуст. Пропуск импорта данных.")
                total_rows = 0
                # Цикл while ниже не выполнится, так как start_row (1) > total_rows (0)

            start_row = 1 # openpyxl использует 1-based индексацию
            while start_row <= total_rows:
                end_row = min(start_row + chunk_size - 1, total_rows)
                logger.debug(f"Обработка строки {start_row} - {end_row} (чанк).")

                raw_data_list = []
                # Используем iter_rows с указанием min_row и max_row для "части"
                for row in sheet.iter_rows(min_row=start_row, max_row=end_row, values_only=False):
                    for cell in row:
                        if cell.value is not None or cell.data_type == 'f':
                            data_item = {
                                "cell_address": cell.coordinate,
                                "value": cell.value,
                            }
                            raw_data_list.append(data_item)

                if not storage.save_sheet_raw_data(sheet_name, raw_data_list):
                    logger.error(f"Не удалось сохранить 'сырые данные' для листа '{sheet_name}' (часть строки {start_row}-{end_row}).")
                    return False

                logger.debug(f"Сохранена часть данных с {start_row} по {end_row} для листа '{sheet_name}'.")

                start_row = end_row + 1 # Переходим к следующей части

        logger.info(f"Импорт 'сырых' данных из '{file_path}' завершён.")
        return True

    except Exception as e:
        logger.error(f"Ошибка при импорте 'сырых' данных из файла '{file_path}': {e}", exc_info=True)
        return False
def import_raw_values_only_from_excel(storage: ProjectDBStorage, file_path: str, options: Optional[Dict[str, Any]] = None) -> bool:
    """
    Импортирует только "сырые" значения (результаты формул и значения ячеек) из Excel-файла в БД проекта.
    Использует openpyxl с data_only=True для получения результатов формул.
    Импортирует частями.

    Args:
        storage: Экземпляр ProjectDBStorage для сохранения данных.
        file_path (str): Путь к Excel-файлу для импорта.
        options (Optional[Dict[str, Any]]): Опции импорта.
            {
                'chunk_size_rows': int, # Количество строк в одной части (по умолчанию 50)
                'sheets': List[str]     # Список имён листов для импорта. Если пуст, все.
            }

    Returns:
        bool: True, если импорт прошёл успешно, иначе False.
    """
    if not storage:
        logger.error("Экземпляр ProjectDBStorage не предоставлен. Невозможно выполнить импорт.")
        return False

    if not os.path.exists(file_path):
        logger.error(f"Excel-файл для импорта не найден: {file_path}")
        return False

    logger.info(f"Начало импорта 'сырых' значений (только результаты) из Excel-файла: {file_path}")

    try:
        # --- НОВОЕ: Обработка ошибки Nested.from_tree ---
        try:
            # Используем data_only=True для получения результатов формул
            workbook = openpyxl.load_workbook(file_path, data_only=True) # <-- КЛЮЧЕВОЕ ИЗМЕНЕНИЕ
        except TypeError as e:
            if "Nested.from_tree() missing 1 required positional argument: 'node'" in str(e):
                logger.error(f"Ошибка openpyxl при открытии файла '{file_path}': {e}")
                logger.warning("Файл может содержать неподдерживаемые структуры (например, pivot-таблицы). Импорт прерван.")
                return False
            else:
                raise # Если это другая ошибка TypeError, пробрасываем её
        # --- КОНЕЦ НОВОГО ---
        logger.debug(f"Книга '{file_path}' успешно открыта с data_only=True.")

        sheets_to_import_orig = options.get('sheets', []) if options else []
        if not sheets_to_import_orig:
            sheets_to_import_orig = workbook.sheetnames

        # Явно приводим элементы к str для устранения ошибки Pylance
        sheets_to_import: List[str] = [str(name) for name in sheets_to_import_orig]

        for sheet_name_orig in sheets_to_import:
            # Убедимся, что имя листа - строка
            sheet_name: str = str(sheet_name_orig)
            if sheet_name not in workbook.sheetnames:
                logger.warning(f"Лист '{sheet_name}' не найден в файле '{file_path}'. Пропущен.")
                continue

            logger.info(f"Импорт 'сырых' значений (только результаты) с листа: {sheet_name}")
            
            # --- НОВОЕ: Гарантируем, что запись о листе существует ---
            # Предполагаем project_id = 1 для MVP
            sheet_id = storage.save_sheet(project_id=1, sheet_name=sheet_name)
            if sheet_id is None:
                logger.error(f"Не удалось создать/получить ID для листа '{sheet_name}'. Пропущен.")
                return False # Возвращаем False при ошибке создания записи о листе
            # --- КОНЕЦ НОВОГО ---

            sheet: Worksheet = workbook[sheet_name]
            total_rows = sheet.max_row
            chunk_size = options.get('chunk_size_rows', 50) if options else 50 # <-- ИЗМЕНЕНО С 100 НА 50
            
            # Исправление: Обработка случая, когда лист пуст (max_row возвращает None)
            if total_rows is None:
                logger.info(f"Лист '{sheet_name}' пуст. Пропуск импорта данных.")
                total_rows = 0
                # Цикл while ниже не выполнится, так как start_row (1) > total_rows (0)

            start_row = 1 # openpyxl использует 1-based индексацию
            while start_row <= total_rows:
                end_row = min(start_row + chunk_size - 1, total_rows)
                logger.debug(f"Обработка строки {start_row} - {end_row} (чанк).")

                raw_data_list = []
                # Используем iter_rows с указанием min_row и max_row для "части"
                for row in sheet.iter_rows(min_row=start_row, max_row=end_row, values_only=False):
                    for cell in row:
                        # При data_only=True cell.value будет содержать результат формулы или значение
                        if cell.value is not None or cell.data_type == 'f': # data_type может быть 'f' даже при data_only=True, если формула была
                            data_item = {
                                "cell_address": cell.coordinate,
                                "value": cell.value, # <-- Это будет результат или значение
                            }
                            raw_data_list.append(data_item)

                if not storage.save_sheet_raw_data(sheet_name, raw_data_list):
                    logger.error(f"Не удалось сохранить 'сырые данные (только результаты)' для листа '{sheet_name}' (часть строки {start_row}-{end_row}).")
                    return False

                logger.debug(f"Сохранена часть данных с {start_row} по {end_row} для листа '{sheet_name}'.")

                start_row = end_row + 1 # Переходим к следующей части

        logger.info(f"Импорт 'сырых' значений (только результаты) из '{file_path}' завершён.")
        return True

    except Exception as e:
        logger.error(f"Ошибка при импорте 'сырых' значений (только результаты) из файла '{file_path}': {e}", exc_info=True)
        return False

# --- ПЕРВАЯ НЕПРАВИЛЬНАЯ ФУНКЦИЯ import_styles_from_excel УДАЛЕНА ---

def import_styles_from_excel(storage: ProjectDBStorage, file_path: str, options: Optional[Dict[str, Any]] = None) -> bool:
    """
    Импортирует только стили из Excel-файла в БД проекта.
    Импортирует частями по строкам.

    Args:
        storage: Экземпляр ProjectDBStorage для сохранения данных.
        file_path (str): Путь к Excel-файлу для импорта.
        options (Optional[Dict[str, Any]]): Опции импорта.
            {
                'chunk_size_rows': int, # Количество строк в одной части (по умолчанию 50)
                'sheets': List[str]     # Список имён листов для импорта. Если пуст, все.
            }

    Returns:
        bool: True, если импорт прошёл успешно, иначе False.
    """
    if not storage:
        logger.error("Экземпляр ProjectDBStorage не предоставлен. Невозможно выполнить импорт.")
        return False

    if not os.path.exists(file_path):
        logger.error(f"Excel-файл для импорта не найден: {file_path}")
        return False

    logger.info(f"Начало импорта стилей из Excel-файла: {file_path}")

    try:
        # --- НОВОЕ: Обработка ошибки Nested.from_tree ---
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=False)
        except TypeError as e:
            if "Nested.from_tree() missing 1 required positional argument: 'node'" in str(e):
                logger.error(f"Ошибка openpyxl при открытии файла '{file_path}': {e}")
                logger.warning("Файл может содержать неподдерживаемые структуры (например, pivot-таблицы). Импорт прерван.")
                return False
            else:
                raise # Если это другая ошибка TypeError, пробрасываем её
        # --- КОНЕЦ НОВОГО ---
        logger.debug(f"Книга '{file_path}' успешно открыта.")

        sheets_to_import_orig = options.get('sheets', []) if options else []
        if not sheets_to_import_orig:
            sheets_to_import_orig = workbook.sheetnames

        # Явно приводим элементы к str для устранения ошибки Pylance
        sheets_to_import: List[str] = [str(name) for name in sheets_to_import_orig]

        import json

        for sheet_name_orig in sheets_to_import:
            # Убедимся, что имя листа - строка
            sheet_name: str = str(sheet_name_orig)
            if sheet_name not in workbook.sheetnames:
                logger.warning(f"Лист '{sheet_name}' не найден в файле '{file_path}'. Пропущен.")
                continue

            logger.info(f"Импорт стилей с листа: {sheet_name}")
            
            # --- НОВОЕ: Гарантируем, что запись о листе существует ---
            # Предполагаем project_id = 1 для MVP
            sheet_id = storage.save_sheet(project_id=1, sheet_name=sheet_name)
            if sheet_id is None:
                logger.error(f"Не удалось создать/получить ID для листа '{sheet_name}'. Пропущен.")
                return False # Возвращаем False при ошибке создания записи о листе
            # --- КОНЕЦ НОВОГО ---

            sheet: Worksheet = workbook[sheet_name]
            total_rows = sheet.max_row
            chunk_size = options.get('chunk_size_rows', 50) if options else 50 # <-- ИЗМЕНЕНО С 100 НА 50
            
            # Исправление: Обработка случая, когда лист пуст (max_row возвращает None)
            if total_rows is None:
                logger.info(f"Лист '{sheet_name}' пуст. Пропуск импорта стилей.")
                total_rows = 0
                # Цикл while ниже не выполнится, так как start_row (1) > total_rows (0)

            start_row = 1 # openpyxl использует 1-based индексацию
            while start_row <= total_rows:
                end_row = min(start_row + chunk_size - 1, total_rows)
                logger.debug(f"Обработка строки {start_row} - {end_row} для стилей (чанк).")

                style_ranges_map: Dict[str, List[str]] = {}

                # Используем iter_rows с указанием min_row и max_row для "части"
                for row in sheet.iter_rows(min_row=start_row, max_row=end_row, values_only=False):
                    for cell in row:
                        style_dict = _serialize_style(cell)
                        if style_dict:
                            style_json = json.dumps(style_dict, sort_keys=True)
                            coord = cell.coordinate

                            if style_json in style_ranges_map:
                                style_ranges_map[style_json].append(coord)
                            else:
                                style_ranges_map[style_json] = [coord]

                styles_to_save = []
                for style_json, cell_addresses in style_ranges_map.items():
                    for address in cell_addresses:
                        styles_to_save.append({
                            "range_address": address,
                            "style_attributes": style_json
                        })

                if not storage.save_sheet_styles(sheet_id, styles_to_save):
                    logger.error(f"Не удалось сохранить стили для листа '{sheet_name}' (часть строки {start_row}-{end_row}).")
                    return False

                logger.debug(f"Сохранена часть стилей с {start_row} по {end_row} для листа '{sheet_name}'.")

                start_row = end_row + 1 # Переходим к следующей части

        logger.info(f"Импорт стилей из '{file_path}' завершён.")
        return True

    except Exception as e:
        logger.error(f"Ошибка при импорте стилей из файла '{file_path}': {e}", exc_info=True)
        return False

def import_charts_from_excel(storage: ProjectDBStorage, file_path: str, options: Optional[Dict[str, Any]] = None) -> bool:
    """
    Импортирует только диаграммы из Excel-файла в БД проекта.
    Импортирует частями по найденным диаграммам (опционально).
    В текущей реализации сохраняет все диаграммы с листа за раз.

    Args:
        storage: Экземпляр ProjectDBStorage для сохранения данных.
        file_path (str): Путь к Excel-файлу для импорта.
        options (Optional[Dict[str, Any]]): Опции импорта.
            {
                'chunk_size_rows': int, # (Игнорируется для диаграмм) Количество строк в одной части (по умолчанию 50)
                'sheets': List[str]     # Список имён листов для импорта. Если пуст, все.
            }

    Returns:
        bool: True, если импорт прошёл успешно, иначе False.
    """
    if not storage:
        logger.error("Экземпляр ProjectDBStorage не предоставлен. Невозможно выполнить импорт.")
        return False

    if not os.path.exists(file_path):
        logger.error(f"Excel-файл для импорта не найден: {file_path}")
        return False

    logger.info(f"Начало импорта диаграмм из Excel-файла: {file_path}")

    try:
        # --- НОВОЕ: Обработка ошибки Nested.from_tree ---
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=False)
        except TypeError as e:
            if "Nested.from_tree() missing 1 required positional argument: 'node'" in str(e):
                logger.error(f"Ошибка openpyxl при открытии файла '{file_path}': {e}")
                logger.warning("Файл может содержать неподдерживаемые структуры (например, pivot-таблицы). Импорт прерван.")
                return False
            else:
                raise # Если это другая ошибка TypeError, пробрасываем её
        # --- КОНЕЦ НОВОГО ---
        logger.debug(f"Книга '{file_path}' успешно открыта.")

        sheets_to_import_orig = options.get('sheets', []) if options else []
        if not sheets_to_import_orig:
            sheets_to_import_orig = workbook.sheetnames

        # Явно приводим элементы к str для устранения ошибки Pylance
        sheets_to_import: List[str] = [str(name) for name in sheets_to_import_orig]

        for sheet_name_orig in sheets_to_import:
            # Убедимся, что имя листа - строка
            sheet_name: str = str(sheet_name_orig)
            if sheet_name not in workbook.sheetnames:
                logger.warning(f"Лист '{sheet_name}' не найден в файле '{file_path}'. Пропущен.")
                continue

            logger.info(f"Импорт диаграмм с листа: {sheet_name}")
            
            # --- НОВОЕ: Гарантируем, что запись о листе существует ---
            # Предполагаем project_id = 1 для MVP
            sheet_id = storage.save_sheet(project_id=1, sheet_name=sheet_name)
            if sheet_id is None:
                logger.error(f"Не удалось создать/получить ID для листа '{sheet_name}'. Пропущен.")
                return False # Возвращаем False при ошибке создания записи о листе
            # --- КОНЕЦ НОВОГО ---

            sheet: Worksheet = workbook[sheet_name]

            charts_list = []
            try:
                # Исправлено: Добавлен # type: ignore[attr-defined] для подавления ошибки Pylance
                charts_sheet = sheet._charts # type: ignore[attr-defined]
                for chart_obj in charts_sheet:
                    chart_data = _serialize_chart(chart_obj)
                    if chart_data:
                        charts_list.append({
                            "chart_data": chart_data
                        })
            except AttributeError as ae:
                logger.warning(f"Не удалось получить доступ к диаграммам листа '{sheet_name}' через _charts: {ae}")
            except Exception as e:
                logger.error(f"Ошибка при извлечении диаграмм с листа '{sheet_name}': {e}", exc_info=True)

            if not storage.save_sheet_charts(sheet_id, charts_list):
                logger.error(f"Не удалось сохранить диаграммы для листа '{sheet_name}' (ID: {sheet_id}).")
                return False

            logger.debug(f"Сохранено {len(charts_list)} записей 'диаграмм' для листа '{sheet_name}'.")

        logger.info(f"Импорт диаграмм из '{file_path}' завершён.")
        return True

    except Exception as e:
        logger.error(f"Ошибка при импорте диаграмм из файла '{file_path}': {e}", exc_info=True)
        return False

def import_formulas_from_excel(storage: ProjectDBStorage, file_path: str, options: Optional[Dict[str, Any]] = None) -> bool:
    """
    Импортирует только формулы из Excel-файла в БД проекта.
    Импортирует частями по строкам.

    Args:
        storage: Экземпляр ProjectDBStorage для сохранения данных.
        file_path (str): Путь к Excel-файлу для импорта.
        options (Optional[Dict[str, Any]]): Опции импорта.
            {
                'chunk_size_rows': int, # Количество строк в одной части (по умолчанию 50)
                'sheets': List[str]     # Список имён листов для импорта. Если пуст, все.
            }

    Returns:
        bool: True, если импорт прошёл успешно, иначе False.
    """
    if not storage:
        logger.error("Экземпляр ProjectDBStorage не предоставлен. Невозможно выполнить импорт.")
        return False

    if not os.path.exists(file_path):
        logger.error(f"Excel-файл для импорта не найден: {file_path}")
        return False

    logger.info(f"Начало импорта формул из Excel-файла: {file_path}")

    try:
        # --- НОВОЕ: Обработка ошибки Nested.from_tree ---
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=False)
        except TypeError as e:
            if "Nested.from_tree() missing 1 required positional argument: 'node'" in str(e):
                logger.error(f"Ошибка openpyxl при открытии файла '{file_path}': {e}")
                logger.warning("Файл может содержать неподдерживаемые структуры (например, pivot-таблицы). Импорт прерван.")
                return False
            else:
                raise # Если это другая ошибка TypeError, пробрасываем её
        # --- КОНЕЦ НОВОГО ---
        logger.debug(f"Книга '{file_path}' успешно открыта.")

        sheets_to_import_orig = options.get('sheets', []) if options else []
        if not sheets_to_import_orig:
            sheets_to_import_orig = workbook.sheetnames

        # Явно приводим элементы к str для устранения ошибки Pylance
        sheets_to_import: List[str] = [str(name) for name in sheets_to_import_orig]

        for sheet_name_orig in sheets_to_import:
            # Убедимся, что имя листа - строка
            sheet_name: str = str(sheet_name_orig)
            if sheet_name not in workbook.sheetnames:
                logger.warning(f"Лист '{sheet_name}' не найден в файле '{file_path}'. Пропущен.")
                continue

            logger.info(f"Импорт формул с листа: {sheet_name}")
            
            # --- НОВОЕ: Гарантируем, что запись о листе существует ---
            # Предполагаем project_id = 1 для MVP
            sheet_id = storage.save_sheet(project_id=1, sheet_name=sheet_name)
            if sheet_id is None:
                logger.error(f"Не удалось создать/получить ID для листа '{sheet_name}'. Пропущен.")
                return False # Возвращаем False при ошибке создания записи о листе
            # --- КОНЕЦ НОВОГО ---

            sheet: Worksheet = workbook[sheet_name]
            total_rows = sheet.max_row
            chunk_size = options.get('chunk_size_rows', 50) if options else 50 # <-- ИЗМЕНЕНО С 100 НА 50
            
            # Исправление: Обработка случая, когда лист пуст (max_row возвращает None)
            if total_rows is None:
                logger.info(f"Лист '{sheet_name}' пуст. Пропуск импорта формул.")
                total_rows = 0
                # Цикл while ниже не выполнится, так как start_row (1) > total_rows (0)

            start_row = 1 # openpyxl использует 1-based индексацию
            while start_row <= total_rows:
                end_row = min(start_row + chunk_size - 1, total_rows)
                logger.debug(f"Обработка строки {start_row} - {end_row} для формул (чанк).")

                formulas_list = []
                # Используем iter_rows с указанием min_row и max_row для "части"
                for row in sheet.iter_rows(min_row=start_row, max_row=end_row, values_only=False):
                    for cell in row:
                        # Исправлено: Проверяем, является ли значение строкой и начинается ли с '='
                        if isinstance(cell.value, str) and cell.value.startswith('='):
                             formulas_list.append({
                                 "cell_address": cell.coordinate,
                                 "formula": cell.value # Сохраняем формулу как есть, включая '='
                             })

                if not storage.save_sheet_formulas(sheet_id, formulas_list):
                    logger.error(f"Не удалось сохранить формулы для листа '{sheet_name}' (часть строки {start_row}-{end_row}).")
                    return False

                logger.debug(f"Сохранена часть формул с {start_row} по {end_row} для листа '{sheet_name}'.")

                start_row = end_row + 1 # Переходим к следующей части

        logger.info(f"Импорт формул из '{file_path}' завершён.")
        return True

    except Exception as e:
        logger.error(f"Ошибка при импорте формул из файла '{file_path}': {e}", exc_info=True)
        return False


# --- Функции для импорта "выборочно" по типам ---

# Заглушка для выборочного импорта. Реализация будет аналогична полному импорту,
# но с фильтрацией по листам/диапазонам, переданным в options.
# Например, options = {'sheets': ['Sheet1'], 'start_row': 1, 'end_row': 100}

def import_raw_data_from_excel_selective(storage: ProjectDBStorage, file_path: str, options: Optional[Dict[str, Any]] = None) -> bool:
    """
    Импортирует "сырые" данные выборочно из Excel-файла в БД проекта.
    Выбор осуществляется на основе опций в `options`.

    Args:
        storage (ProjectDBStorage): Экземпляр ProjectDBStorage для сохранения данных.
        file_path (str): Путь к Excel-файлу для импорта.
        options (Optional[Dict[str, Any]]): Опции импорта.
            {
                'sheets': List[str],     # Список имён листов для импорта.
                'chunk_size_rows': int, # Количество строк в одной части (по умолчанию 50)
                # Другие опции в будущем...
            }

    Returns:
        bool: True, если импорт прошёл успешно, иначе False.
    """
    logger.info(f"Начало выборочного импорта 'сырых' данных из '{file_path}' с опциями {options}.")
    # Для выборочного импорта "сырых данных" просто передаём options в основную функцию
    # Основная функция уже умеет обрабатывать options['sheets']
    return import_raw_data_from_excel(storage, file_path, options)
def import_styles_from_excel_selective(storage: ProjectDBStorage, file_path: str, options: Optional[Dict[str, Any]] = None) -> bool:
    """
    Импортирует стили выборочно из Excel-файла в БД проекта.

    Args:
        storage (ProjectDBStorage): Экземпляр ProjectDBStorage для сохранения данных.
        file_path (str): Путь к Excel-файлу для импорта.
        options (Optional[Dict[str, Any]]): Опции импорта.
            {
                'sheets': List[str],     # Список имён листов для импорта.
                'chunk_size_rows': int, # Количество строк в одной части (по умолчанию 50)
                # Другие опции в будущем...
            }

    Returns:
        bool: True, если импорт прошёл успешно, иначе False.
    """
    logger.info(f"Начало выборочного импорта 'стилей' из '{file_path}' с опциями {options}.")
    return import_styles_from_excel(storage, file_path, options)
def import_charts_from_excel_selective(storage: ProjectDBStorage, file_path: str, options: Optional[Dict[str, Any]] = None) -> bool:
    """
    Импортирует диаграммы выборочно из Excel-файла в БД проекта.

    Args:
        storage (ProjectDBStorage): Экземпляр ProjectDBStorage для сохранения данных.
        file_path (str): Путь к Excel-файлу для импорта.
        options (Optional[Dict[str, Any]]): Опции импорта.
            {
                'sheets': List[str],     # Список имён листов для импорта.
                # 'chunk_size_charts': int, # Количество диаграмм в одной части (по умолчанию len(all_charts_on_sheet))
                # Другие опции в будущем...
            }

    Returns:
        bool: True, если импорт прошёл успешно, иначе False.
    """
    logger.info(f"Начало выборочного импорта 'диаграмм' из '{file_path}' с опциями {options}.")
    return import_charts_from_excel(storage, file_path, options)
def import_formulas_from_excel_selective(storage: ProjectDBStorage, file_path: str, options: Optional[Dict[str, Any]] = None) -> bool:
    """
    Импортирует формулы выборочно из Excel-файла в БД проекта.

    Args:
        storage (ProjectDBStorage): Экземпляр ProjectDBStorage для сохранения данных.
        file_path (str): Путь к Excel-файлу для импорта.
        options (Optional[Dict[str, Any]]): Опции импорта.
            {
                'sheets': List[str],     # Список имён листов для импорта.
                'chunk_size_rows': int, # Количество строк в одной части (по умолчанию 50)
                # Другие опции в будущем...
            }

    Returns:
        bool: True, если импорт прошёл успешно, иначе False.
    """
    logger.info(f"Начало выборочного импорта 'формул' из '{file_path}' с опциями {options}.")
    return import_formulas_from_excel(storage, file_path, options)
def import_all_data_from_excel_selective(storage: ProjectDBStorage, file_path: str, options: Optional[Dict[str, Any]] = None) -> bool:
    """
    Импортирует все типы данных выборочно из Excel-файла.
    Использует чанкирование через обновлённые функции импорта "всё".

    Args:
        storage (ProjectDBStorage): Экземпляр ProjectDBStorage для сохранения данных.
        file_path (str): Путь к Excel-файлу для импорта.
        options (Optional[Dict[str, Any]]): Опции импорта.
            {
                'sheets': List[str],     # Список имён листов для импорта.
                'chunk_size_rows': int, # Количество строк в одной части (по умолчанию 50)
                # 'chunk_size_charts': int, # Количество диаграмм в одной части (по умолчанию len(all_charts_on_sheet))
                # Другие опции в будущем...
            }

    Returns:
        bool: True, если импорт прошёл успешно, иначе False.
    """
    logger.info(f"Начало выборочного импорта всех данных из '{file_path}' с опциями {options}.")

    # Вызываем каждую функцию импорта "всё" с теми же options
    # Они сами обработают options['sheets']
    success_raw = import_raw_data_from_excel_selective(storage, file_path, options)
    success_styles = import_styles_from_excel_selective(storage, file_path, options)
    success_charts = import_charts_from_excel_selective(storage, file_path, options)
    success_formulas = import_formulas_from_excel_selective(storage, file_path, options)

    overall_success = success_raw and success_styles and success_charts and success_formulas

    if overall_success:
        logger.info(f"Выборочный импорт всех данных из '{file_path}' завершён успешно.")
    else:
        logger.error(f"Выборочный импорт всех данных из '{file_path}' завершён с ошибками.")

    return overall_success


# --- Функции для импорта "частями" по типам ---

# Заглушка для импорта частями. Реализация будет разбивать большой файл
# на части и вызывать соответствующую функцию импорта для каждой части.
# Например, импортировать по 1000 строк за раз.

def import_raw_data_from_excel_in_chunks(storage: ProjectDBStorage, file_path: str, chunk_options: Dict[str, Any]) -> bool:
    """
    Импортирует "сырые" данные (значения ячеек) частями из Excel-файла в БД проекта.
    Использует openpyxl и параметр 'chunk_size_rows' из chunk_options.

    Args:
        storage: Экземпляр ProjectDBStorage для сохранения данных.
        file_path (str): Путь к Excel-файлу для импорта.
        chunk_options (Dict[str, Any]): Опции для разбиения на части.
            {
                'chunk_size_rows': int, # Количество строк в одной части (по умолчанию 50)
                'sheets': List[str]     # Список имён листов для импорта. Если пуст, все.
            }

    Returns:
        bool: True, если импорт прошёл успешно, иначе False.
    """
    # Теперь основная логика импорта "всё" использует чанки.
    # Эта функция может быть упразднена или стать алиасом для import_raw_data_from_excel.
    logger.warning("Функция import_raw_data_from_excel_in_chunks устарела. Используйте import_raw_data_from_excel с опцией chunk_size_rows.")
    return import_raw_data_from_excel(storage, file_path, chunk_options)
def import_styles_from_excel_in_chunks(storage: ProjectDBStorage, file_path: str, chunk_options: Dict[str, Any]) -> bool:
    """
    Импортирует стили частями из Excel-файла в БД проекта.
    """
    logger.warning("Функция import_styles_from_excel_in_chunks устарела. Используйте import_styles_from_excel с опцией chunk_size_rows.")
    return import_styles_from_excel(storage, file_path, chunk_options)
def import_charts_from_excel_in_chunks(storage: ProjectDBStorage, file_path: str, chunk_options: Dict[str, Any]) -> bool:
    """
    Импортирует диаграммы частями из Excel-файла в БД проекта.
    """
    logger.warning("Функция import_charts_from_excel_in_chunks устарела. Используйте import_charts_from_excel с опцией chunk_size_charts.")
    return import_charts_from_excel(storage, file_path, chunk_options)
def import_formulas_from_excel_in_chunks(storage: ProjectDBStorage, file_path: str, chunk_options: Dict[str, Any]) -> bool:
    """
    Импортирует формулы частями из Excel-файла в БД проекта.
    """
    logger.warning("Функция import_formulas_from_excel_in_chunks устарела. Используйте import_formulas_from_excel с опцией chunk_size_rows.")
    return import_formulas_from_excel(storage, file_path, chunk_options)
def import_all_data_from_excel_chunks(storage: ProjectDBStorage, file_path: str, chunk_options: Dict[str, Any]) -> bool:
    """
    Импортирует все типы данных частями.
    """
    logger.warning("Функция import_all_data_from_excel_chunks устарела. Используйте import_all_data_from_excel с опцией chunk_size_rows.")
    # Для совместимости, передаём chunk_options в import_all_data_from_excel
    # Предполагаем, что import_all_data_from_excel также будет использовать chunk_size_rows
    return import_all_data_from_excel(storage, file_path, chunk_options)


# --- УПРАЗДНЕНИЕ РЕЖИМА АВТО ---
# Функция import_auto_data_from_excel упразднена.
# def import_auto_data_from_excel(...) -> bool:
#     # Эта функция больше не существует.
#     pass

# --- ОБНОВЛЕНИЕ ФУНКЦИИ ИМПОРТА ВСЕГО ---

def import_all_data_from_excel(storage: ProjectDBStorage, file_path: str, options: Optional[Dict[str, Any]] = None) -> bool:
    """
    Импортирует все поддерживаемые типы данных (сырые, стили, диаграммы, формулы) из Excel-файла.
    Использует чанкирование через обновлённые функции импорта "всё".

    Args:
        storage: Экземпляр ProjectDBStorage для сохранения данных.
        file_path (str): Путь к Excel-файлу для импорта.
        options (Optional[Dict[str, Any]]): Опции импорта.
            {
                'chunk_size_rows': int, # Количество строк в одной части (по умолчанию 50)
                'chunk_size_charts': int, # Количество диаграмм в одной части (по умолчанию len(all_charts_on_sheet))
                'sheets': List[str]     # Список имён листов для импорта. Если пуст, все.
            }

    Returns:
        bool: True, если импорт прошёл успешно, иначе False.
    """
    if not storage:
        logger.error("Экземпляр ProjectDBStorage не предоставлен. Невозможно выполнить импорт.")
        return False

    if not os.path.exists(file_path):
        logger.error(f"Excel-файл для импорта не найден: {file_path}")
        return False

    logger.info(f"Начало импорта всех данных из Excel-файла: {file_path}")

    # Вызываем каждую функцию импорта "всё" с одинаковыми options
    success_raw = import_raw_data_from_excel(storage, file_path, options)
    success_styles = import_styles_from_excel(storage, file_path, options)
    success_charts = import_charts_from_excel(storage, file_path, options)
    success_formulas = import_formulas_from_excel(storage, file_path, options)

    overall_success = success_raw and success_styles and success_charts and success_formulas

    if overall_success:
        logger.info(f"Импорт всех данных из '{file_path}' завершён успешно.")
    else:
        logger.error(f"Импорт всех данных из '{file_path}' завершён с ошибками.")

    return overall_success