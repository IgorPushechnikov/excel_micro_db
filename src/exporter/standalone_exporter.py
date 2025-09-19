# src/exporter/standalone_exporter.py
"""
Самостоятельный скрипт экспорта проекта Excel Micro DB в файл Excel (.xlsx).
Объединяет экспорт данных, стилей и диаграмм в один проход с помощью openpyxl.
Предназначен для отладки и обеспечения корректного экспорта.
"""

import sys
import os
from pathlib import Path
import sqlite3
import logging
from typing import Dict, Any, List, Optional, Tuple, Union
import hashlib # Для генерации имен стилей

# Импорты openpyxl
from openpyxl import Workbook
from openpyxl.workbook.workbook import Workbook as OpenpyxlWorkbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.cell_range import CellRange
# Импорты для стилей
from openpyxl.styles import (
    Font, Fill, Border, PatternFill, Side, Alignment, Protection, NamedStyle, Color
)
# Импорты для диаграмм
from openpyxl.chart import BarChart, LineChart, PieChart, Reference # Добавьте другие типы по необходимости
from openpyxl.chart.label import DataLabelList

# --- Настройка логирования ---
# Этот логгер будет выводить информацию непосредственно в консоль/терминал,
# что удобно для отладки во время тестов.
logger = logging.getLogger("standalone_exporter")
logger.setLevel(logging.DEBUG)
if not logger.handlers:
    handler = logging.StreamHandler(sys.stdout)
    handler.setLevel(logging.DEBUG)
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    handler.setFormatter(formatter)
    logger.addHandler(handler)
# --- Конец настройки логирования ---

def _add_project_root_to_path_if_needed():
    """Добавляет корень проекта в sys.path, если модуль src не найден."""
    try:
        # Попробуем импортировать, чтобы проверить, доступен ли он
        import src
        logger.debug("[ЭКСПОРТ] Модуль 'src' уже доступен в sys.path.")
    except ImportError:
        logger.debug("[ЭКСПОРТ] Модуль 'src' не найден. Попытка добавить корень проекта в sys.path.")
        # Определяем корень проекта как директорию, содержащую 'src'
        # Используем __file__ этого скрипта для определения пути
        current_script_path = Path(__file__).resolve()
        # Поднимаемся вверх по дереву каталогов, пока не найдем 'src' или не дойдем до корня
        potential_root = current_script_path.parent
        max_levels_up = 10 # Ограничение на количество уровней вверх для поиска
        root_found = False
        for _ in range(max_levels_up):
            if (potential_root / 'src').is_dir():
                project_root_candidate = potential_root
                root_found = True
                break
            potential_root = potential_root.parent

        if root_found:
            project_root_str = str(project_root_candidate)
            if project_root_str not in sys.path:
                sys.path.insert(0, project_root_str)
                logger.debug(f"[ЭКСПОРТ] Корень проекта '{project_root_str}' добавлен в sys.path.")
            else:
                logger.debug(f"[ЭКСПОРТ] Корень проекта '{project_root_str}' уже в sys.path.")
        else:
            logger.warning("[ЭКСПОРТ] Не удалось автоматически определить корень проекта для добавления в sys.path.")

# Выполняем проверку и добавление пути при импорте модуля
_add_project_root_to_path_if_needed()

# --- Функции для работы со стилями (перенесены и адаптированы из style_exporter.py) ---

def _create_openpyxl_font_from_db_row(font_row: sqlite3.Row) -> Optional[Font]:
    """Создает объект Font openpyxl из строки БД."""
    if not font_row:
        return None
    try:
        font_kwargs = {}
        # Используем .keys() и обращение по индексу или ключу для sqlite3.Row
        if 'name' in font_row.keys() and font_row['name']:
            font_kwargs['name'] = font_row['name']
        if 'sz' in font_row.keys() and font_row['sz'] is not None:
            font_kwargs['sz'] = float(font_row['sz'])
        font_kwargs['bold'] = bool(font_row['b']) if 'b' in font_row.keys() and font_row['b'] is not None else False
        font_kwargs['italic'] = bool(font_row['i']) if 'i' in font_row.keys() and font_row['i'] is not None else False
        if 'u' in font_row.keys() and font_row['u']:
            font_kwargs['underline'] = font_row['u']
        font_kwargs['strike'] = bool(font_row['strike']) if 'strike' in font_row.keys() and font_row['strike'] is not None else False
        if 'vert_align' in font_row.keys() and font_row['vert_align']:
            font_kwargs['vertAlign'] = font_row['vert_align']
        if 'scheme' in font_row.keys() and font_row['scheme']:
            font_kwargs['scheme'] = font_row['scheme']

        # Обработка цвета шрифта
        color_kwargs = {}
        if 'color' in font_row.keys() and font_row['color']:
            color_kwargs['rgb'] = font_row['color']
        elif 'color_theme' in font_row.keys() and font_row['color_theme'] is not None:
            color_kwargs['theme'] = int(font_row['color_theme'])
            if 'color_tint' in font_row.keys() and font_row['color_tint'] is not None:
                color_kwargs['tint'] = float(font_row['color_tint'])
        if color_kwargs:
            try:
                font_kwargs['color'] = Color(**color_kwargs)
            except Exception as e:
                logger.warning(f"[СТИЛЬ] Ошибка создания цвета шрифта: {e}")

        logger.debug(f"[СТИЛЬ] Создан Font: {font_kwargs}")
        return Font(**font_kwargs)
    except Exception as e:
        logger.error(f"[СТИЛЬ] Ошибка создания Font из строки БД {dict(font_row) if hasattr(font_row, 'keys') else 'N/A'}: {e}")
        return None

def _create_openpyxl_fill_from_db_row(fill_row: sqlite3.Row) -> Optional[Fill]:
    """Создает объект Fill openpyxl из строки БД."""
    if not fill_row or ('pattern_type' not in fill_row.keys()) or not fill_row['pattern_type']:
        return None
    try:
        fill_kwargs = {}
        fill_kwargs['patternType'] = fill_row['pattern_type']

        # Обработка цвета переднего плана
        fg_color_kwargs = {}
        if 'fg_color' in fill_row.keys() and fill_row['fg_color']:
            fg_color_kwargs['rgb'] = fill_row['fg_color']
        elif 'fg_color_theme' in fill_row.keys() and fill_row['fg_color_theme'] is not None:
            fg_color_kwargs['theme'] = int(fill_row['fg_color_theme'])
            if 'fg_color_tint' in fill_row.keys() and fill_row['fg_color_tint'] is not None:
                fg_color_kwargs['tint'] = float(fill_row['fg_color_tint'])
        if fg_color_kwargs:
            try:
                fill_kwargs['fgColor'] = Color(**fg_color_kwargs)
            except Exception as e:
                logger.warning(f"[СТИЛЬ] Ошибка создания fgColor для Fill: {e}")

        # Обработка цвета фона
        bg_color_kwargs = {}
        if 'bg_color' in fill_row.keys() and fill_row['bg_color']:
            bg_color_kwargs['rgb'] = fill_row['bg_color']
        elif 'bg_color_theme' in fill_row.keys() and fill_row['bg_color_theme'] is not None:
            bg_color_kwargs['theme'] = int(fill_row['bg_color_theme'])
            if 'bg_color_tint' in fill_row.keys() and fill_row['bg_color_tint'] is not None:
                bg_color_kwargs['tint'] = float(fill_row['bg_color_tint'])
        if bg_color_kwargs:
            try:
                fill_kwargs['bgColor'] = Color(**bg_color_kwargs)
            except Exception as e:
                logger.warning(f"[СТИЛЬ] Ошибка создания bgColor для Fill: {e}")

        fill_obj = PatternFill(**fill_kwargs)
        logger.debug(f"[СТИЛЬ] Создан Fill: {fill_obj}")
        return fill_obj
    except Exception as e:
        logger.error(f"[СТИЛЬ] Ошибка создания Fill из строки БД {dict(fill_row) if hasattr(fill_row, 'keys') else 'N/A'}: {e}")
        return None

def _create_openpyxl_side_from_attrs(style: Optional[str], color: Optional[str]) -> Optional[Side]:
    """Создает объект Side openpyxl из стиля и цвета."""
    if not style and not color:
        return None
    try:
        side_kwargs = {}
        if style:
            side_kwargs['style'] = style
        if color:
            try:
                side_kwargs['color'] = Color(rgb=color)
            except Exception as e:
                logger.warning(f"[СТИЛЬ] Ошибка создания цвета для Side '{color}': {e}")
        side_obj = Side(**side_kwargs)
        logger.debug(f"[СТИЛЬ] Создан Side: {side_obj}")
        return side_obj
    except Exception as e:
        logger.error(f"[СТИЛЬ] Ошибка создания Side из style='{style}', color='{color}': {e}")
        return None

def _create_openpyxl_border_from_db_row(border_row: sqlite3.Row) -> Border:
    """Создает объект Border openpyxl из строки БД."""
    try:
        sides = {}
        for side_name in ['left', 'right', 'top', 'bottom']:
            style_key = f'{side_name}_style'
            color_key = f'{side_name}_color'
            style = border_row[style_key] if style_key in border_row.keys() else None
            color = border_row[color_key] if color_key in border_row.keys() else None
            side_obj = _create_openpyxl_side_from_attrs(style, color)
            if side_obj:
                sides[side_name] = side_obj

        # Диагональ обрабатывается отдельно, если нужна
        diagonal_style = border_row.get('diagonal_style') if 'diagonal_style' in border_row.keys() else None
        diagonal_color = border_row.get('diagonal_color') if 'diagonal_color' in border_row.keys() else None
        diagonal_side = _create_openpyxl_side_from_attrs(diagonal_style, diagonal_color)
        if diagonal_side:
            sides['diagonal'] = diagonal_side

        # Используем get с дефолтами None для булевых значений
        diagonal_up = bool(border_row.get('diagonal_up')) if border_row.get('diagonal_up') is not None else None
        diagonal_down = bool(border_row.get('diagonal_down')) if border_row.get('diagonal_down') is not None else None
        outline = bool(border_row.get('outline')) if border_row.get('outline') is not None else None

        border_obj = Border(
            left=sides.get('left'),
            right=sides.get('right'),
            top=sides.get('top'),
            bottom=sides.get('bottom'),
            diagonal=sides.get('diagonal'),
            diagonalUp=diagonal_up,
            diagonalDown=diagonal_down,
            outline=outline
        )
        logger.debug(f"[СТИЛЬ] Создан Border: {border_obj}")
        return border_obj
    except Exception as e:
        logger.error(f"[СТИЛЬ] Ошибка создания Border из строки БД {dict(border_row) if hasattr(border_row, 'keys') else 'N/A'}: {e}")
        return Border() # Возвращаем пустую границу

def _create_openpyxl_alignment_from_db_row(align_row: sqlite3.Row) -> Alignment:
    """Создает объект Alignment openpyxl из строки БД."""
    try:
        align_kwargs = {}
        if 'horizontal' in align_row.keys() and align_row['horizontal']:
            align_kwargs['horizontal'] = align_row['horizontal']
        if 'vertical' in align_row.keys() and align_row['vertical']:
            align_kwargs['vertical'] = align_row['vertical']
        if 'text_rotation' in align_row.keys() and align_row['text_rotation'] is not None:
            align_kwargs['textRotation'] = int(align_row['text_rotation'])
        align_kwargs['wrapText'] = bool(align_row['wrap_text']) if 'wrap_text' in align_row.keys() and align_row['wrap_text'] is not None else False
        align_kwargs['shrinkToFit'] = bool(align_row['shrink_to_fit']) if 'shrink_to_fit' in align_row.keys() and align_row['shrink_to_fit'] is not None else False
        if 'indent' in align_row.keys() and align_row['indent'] is not None:
            align_kwargs['indent'] = int(align_row['indent'])
        # relativeIndent, justifyLastLine, readingOrder, textDirection - можно добавить при необходимости
        alignment_obj = Alignment(**align_kwargs)
        logger.debug(f"[СТИЛЬ] Создан Alignment: {alignment_obj}")
        return alignment_obj
    except Exception as e:
        logger.error(f"[СТИЛЬ] Ошибка создания Alignment из строки БД {dict(align_row) if hasattr(align_row, 'keys') else 'N/A'}: {e}")
        return Alignment()

def _create_openpyxl_protection_from_db_row(prot_row: sqlite3.Row) -> Protection:
    """Создает объект Protection openpyxl из строки БД."""
    try:
        prot_kwargs = {}
        # По умолчанию в Excel ячейки заблокированы (locked=True)
        prot_kwargs['locked'] = bool(prot_row['locked']) if 'locked' in prot_row.keys() and prot_row['locked'] is not None else True
        prot_kwargs['hidden'] = bool(prot_row['hidden']) if 'hidden' in prot_row.keys() and prot_row['hidden'] is not None else False
        protection_obj = Protection(**prot_kwargs)
        logger.debug(f"[СТИЛЬ] Создан Protection: {protection_obj}")
        return protection_obj
    except Exception as e:
        logger.error(f"[СТИЛЬ] Ошибка создания Protection из строки БД {dict(prot_row) if hasattr(prot_row, 'keys') else 'N/A'}: {e}")
        return Protection()

def _generate_style_name(style_attrs: Dict[str, Any]) -> str:
    """Генерирует уникальное имя стиля на основе его атрибутов."""
    # Создаем строку атрибутов, отсортированную по ключам, для детерминированного хэша
    # Фильтруем None, так как они не влияют на стиль, но могут мешать хэшированию
    filtered_attrs = {k: v for k, v in style_attrs.items() if v is not None}
    attr_str = str(sorted(filtered_attrs.items()))
    # Генерируем хэш
    style_hash = hashlib.md5(attr_str.encode('utf-8')).hexdigest()[:10]
    return f"Style_{style_hash}"

def _create_named_style_from_db_joined_data(style_data: Dict[str, Any]) -> Optional[Tuple[str, NamedStyle]]:
    """
    Создает именованный стиль openpyxl из объединенных данных стиля из БД.
    Args:
        style_data (Dict[str, Any]): Словарь с данными стиля, объединенными из разных таблиц.
    Returns:
        Optional[Tuple[str, NamedStyle]]: Кортеж (имя_стиля, объект_NamedStyle) или None.
    """
    try:
        style_name = _generate_style_name(style_data)
        logger.debug(f"[СТИЛЬ] Создание именованного стиля '{style_name}'")

        named_style = NamedStyle(name=style_name)

        # Создаем и применяем компоненты
        # 1. Font
        font_attrs = {k.split('_', 1)[1]: v for k, v in style_data.items() if k.startswith('font_') and v is not None}
        if font_attrs:
             # Создаем "фальшивую" строку sqlite3.Row для совместимости с существующей функцией
             # Это простой объект, имитирующий поведение sqlite3.Row для чтения
             class FakeRow:
                 def __init__(self, data):
                     self._data = data
                 def keys(self):
                     return self._data.keys()
                 def __getitem__(self, key):
                     return self._data.get(key)
             font_row = FakeRow(font_attrs)
             font_obj = _create_openpyxl_font_from_db_row(font_row)
             if font_obj:
                 named_style.font = font_obj
                 logger.debug(f"[СТИЛЬ] Font применен к стилю '{style_name}'")

        # 2. Fill
        fill_attrs = {k.split('_', 1)[1]: v for k, v in style_data.items() if k.startswith('fill_') and v is not None}
        if fill_attrs:
             class FakeRow:
                 def __init__(self, data):
                     self._data = data
                 def keys(self):
                     return self._data.keys()
                 def __getitem__(self, key):
                     return self._data.get(key)
                 def get(self, key, default=None):
                     return self._data.get(key, default)
             fill_row = FakeRow(fill_attrs)
             fill_obj = _create_openpyxl_fill_from_db_row(fill_row)
             if fill_obj:
                 named_style.fill = fill_obj
                 logger.debug(f"[СТИЛЬ] Fill применен к стилю '{style_name}'")

        # 3. Border
        border_attrs = {k.split('_', 1)[1]: v for k, v in style_data.items() if k.startswith('border_') and v is not None}
        if border_attrs:
             class FakeRow:
                 def __init__(self, data):
                     self._data = data
                 def keys(self):
                     return self._data.keys()
                 def __getitem__(self, key):
                     return self._data.get(key)
                 def get(self, key, default=None):
                     return self._data.get(key, default)
             border_row = FakeRow(border_attrs)
             border_obj = _create_openpyxl_border_from_db_row(border_row)
             named_style.border = border_obj
             logger.debug(f"[СТИЛЬ] Border применен к стилю '{style_name}'")

        # 4. Alignment
        align_attrs = {k.split('_', 1)[1]: v for k, v in style_data.items() if k.startswith('alignment_') and v is not None}
        if align_attrs:
             class FakeRow:
                 def __init__(self, data):
                     self._data = data
                 def keys(self):
                     return self._data.keys()
                 def __getitem__(self, key):
                     return self._data.get(key)
                 def get(self, key, default=None):
                     return self._data.get(key, default)
             align_row = FakeRow(align_attrs)
             align_obj = _create_openpyxl_alignment_from_db_row(align_row)
             named_style.alignment = align_obj
             logger.debug(f"[СТИЛЬ] Alignment применен к стилю '{style_name}'")

        # 5. Protection
        prot_attrs = {k.split('_', 1)[1]: v for k, v in style_data.items() if k.startswith('protection_') and v is not None}
        if prot_attrs:
             class FakeRow:
                 def __init__(self, data):
                     self._data = data
                 def keys(self):
                     return self._data.keys()
                 def __getitem__(self, key):
                     return self._data.get(key)
                 def get(self, key, default=None):
                     return self._data.get(key, default)
             prot_row = FakeRow(prot_attrs)
             prot_obj = _create_openpyxl_protection_from_db_row(prot_row)
             named_style.protection = prot_obj
             logger.debug(f"[СТИЛЬ] Protection применен к стилю '{style_name}'")

        logger.info(f"[СТИЛЬ] Именованный стиль '{style_name}' успешно создан.")
        return style_name, named_style

    except Exception as e:
        logger.error(f"[СТИЛЬ] Ошибка создания именованного стиля из данных {style_data}: {e}", exc_info=True)
        return None

# --- Функции для экспорта данных ---

def export_sheet_data_and_formulas(ws: Worksheet, sheet_data: Dict[str, Any]) -> bool:
    """Экспортирует данные и формулы на лист."""
    try:
        logger.info(f"[ДАННЫЕ] === Начало экспорта данных и формул для листа '{ws.title}' ===")
        
        # 1. Создание структуры (заголовки)
        structure = sheet_data.get("structure", [])
        logger.debug(f"[ДАННЫЕ] Экспорт структуры: {len(structure)} колонок")
        for col_info in structure:
            col_idx = col_info.get("column_index", 0)
            if col_idx > 0: # Excel колонки начинаются с 1
                header = col_info.get("column_name", f"Col{col_idx}")
                cell = ws.cell(row=1, column=col_idx, value=header)
                logger.debug(f"[ДАННЫЕ] Запись заголовка '{header}' в ячейку {cell.coordinate}")

        # 2. Заполнение данными
        raw_data = sheet_data.get("raw_data", [])
        logger.debug(f"[ДАННЫЕ] Экспорт данных: {len(raw_data)} строк")
        for row_idx, row_data in enumerate(raw_data, start=2): # Начинаем со второй строки
             # Предполагаем, что row_data это список значений
             for col_idx, cell_value in enumerate(row_data, start=1):
                 cell = ws.cell(row=row_idx, column=col_idx)
                 # Обработка формул
                 if isinstance(cell_value, str) and cell_value.startswith('='):
                     logger.debug(f"[ДАННЫЕ] Запись формулы '{cell_value}' в ячейку {cell.coordinate}")
                     cell.value = cell_value # openpyxl автоматически обработает формулу
                 else:
                     logger.debug(f"[ДАННЫЕ] Запись значения '{cell_value}' в ячейку {cell.coordinate}")
                     cell.value = cell_value
        
        logger.info(f"[ДАННЫЕ] === Завершен экспорт данных и формул для листа '{ws.title}' ===")
        return True
    except Exception as e:
        logger.error(f"[ДАННЫЕ] Ошибка экспорта данных и формул для листа '{ws.title}': {e}", exc_info=True)
        return False

# --- Функции для экспорта стилей ---

def export_sheet_styles(wb: OpenpyxlWorkbook, ws: Worksheet, styled_ranges_data: List[Dict[str, Any]]) -> bool:
    """Экспортирует стили на лист."""
    try:
        logger.info(f"[СТИЛЬ] === Начало экспорта стилей для листа '{ws.title}' ===")
        logger.debug(f"[СТИЛЬ] Получено {len(styled_ranges_data)} записей о стилях.")

        if not styled_ranges_data:
            logger.info("[СТИЛЬ] Нет стилей для экспорта.")
            return True

        applied_styles_count = 0
        # Словарь для кэширования уже созданных стилей
        created_styles_cache: Dict[str, NamedStyle] = {}

        for i, style_range_info in enumerate(styled_ranges_data):
            logger.debug(f"[СТИЛЬ] Обработка стиля {i+1}/{len(styled_ranges_data)}: {style_range_info}")
            try:
                range_addr = style_range_info.get("range_address", "")
                if not range_addr:
                     logger.warning(f"[СТИЛЬ] Пропущена запись стиля {i+1} из-за отсутствия 'range_address': {style_range_info}")
                     continue

                # Генерируем имя стиля на основе его атрибутов
                # Исключаем range_address из хеширования
                attrs_for_hash = {k: v for k, v in style_range_info.items() if k != "range_address" and v is not None}
                style_name = _generate_style_name(attrs_for_hash)
                logger.debug(f"[СТИЛЬ] Сгенерировано имя стиля: {style_name}")

                # Проверяем, создан ли стиль
                named_style = None
                if style_name in created_styles_cache:
                    named_style = created_styles_cache[style_name]
                    logger.debug(f"[СТИЛЬ] Стиль '{style_name}' найден в кэше.")
                else:
                    # Создаем именованный стиль
                    style_creation_result = _create_named_style_from_db_joined_data(style_range_info)
                    if style_creation_result:
                        created_style_name, created_named_style = style_creation_result
                        # Добавляем стиль в книгу
                        try:
                            wb.add_named_style(created_named_style)
                            logger.debug(f"[СТИЛЬ] Добавлен новый именованный стиль '{created_style_name}' в книгу.")
                            created_styles_cache[style_name] = created_named_style
                            named_style = created_named_style
                        except Exception as add_style_e:
                            if "already exists" in str(add_style_e):
                                logger.info(f"[СТИЛЬ] Стиль '{created_style_name}' уже существует в книге (ошибка при добавлении).")
                                # Пытаемся получить его из книги
                                if created_style_name in wb._named_styles:
                                    # wb._named_styles это список имен или объектов?
                                    # В новых версиях это словарь {name: NamedStyle}
                                    # named_style = wb._named_styles[created_style_name]
                                    # Более безопасный способ:
                                    try:
                                        # Индекс по имени в списке named_styles
                                        style_index = wb.named_styles.index(created_style_name)
                                        named_style = wb._named_styles[style_index]
                                    except (ValueError, IndexError):
                                        logger.warning(f"[СТИЛЬ] Стиль '{created_style_name}' не найден в внутреннем списке _named_styles.")
                                        # Пробуем создать временный стиль или пропустить
                                        named_style = created_named_style # Используем созданный объект
                                else:
                                    logger.error(f"[СТИЛЬ] Стиль '{created_style_name}' не найден в книге после ошибки добавления. Пропущен.")
                                    continue
                            else:
                                logger.error(f"[СТИЛЬ] Ошибка добавления стиля '{created_style_name}' в книгу: {add_style_e}")
                                continue
                    else:
                        logger.error(f"[СТИЛЬ] Не удалось создать именованный стиль для {attrs_for_hash}")
                        continue

                # Применяем стиль к диапазону на листе
                try:
                    cell_range = ws[range_addr]
                    cells_to_style: List[Cell] = []
                    if isinstance(cell_range, Cell):
                        cells_to_style = [cell_range]
                    elif hasattr(cell_range, '__iter__'):
                        try:
                            for item in cell_range:
                                if isinstance(item, Cell):
                                    cells_to_style.append(item)
                                elif hasattr(item, '__iter__'):
                                    # Это CellRange, итерируемся по ячейкам внутри
                                    for cell in item:
                                        cells_to_style.append(cell)
                        except TypeError:
                            # Если cell_range не итерируемый объект, который мы ожидаем
                            logger.warning(f"[СТИЛЬ] Неожиданный тип для диапазона '{range_addr}': {type(cell_range)}. Пропущен.")
                            continue

                    logger.debug(f"[СТИЛЬ] Применение стиля '{named_style.name}' к {len(cells_to_style)} ячейкам в диапазоне '{range_addr}'.")
                    applied_to_count = 0
                    for cell in cells_to_style:
                        try:
                            # Проверяем, что стиль существует в книге перед применением
                            # Более надежная проверка через внутренний словарь
                            if named_style.name in wb._named_styles:
                                cell.style = named_style.name # Применяем по имени
                                applied_to_count += 1
                                logger.debug(f"[СТИЛЬ] Стиль '{named_style.name}' применен к ячейке {cell.coordinate}.")
                            else:
                                logger.warning(f"[СТИЛЬ] Стиль '{named_style.name}' не найден в книге при проверке перед применением к {cell.coordinate}.")
                        except Exception as apply_cell_style_e:
                            logger.error(f"[СТИЛЬ] Ошибка применения стиля '{named_style.name}' к ячейке {cell.coordinate}: {apply_cell_style_e}")

                    logger.debug(f"[СТИЛЬ] Стиль '{style_name}' применен к {applied_to_count}/{len(cells_to_style)} ячейкам в диапазоне '{range_addr}'.")
                    if applied_to_count > 0:
                        applied_styles_count += 1
                    logger.debug(f"[СТИЛЬ] Обработка диапазона '{range_addr}' завершена.")
                except Exception as apply_range_e:
                    logger.error(f"[СТИЛЬ] Ошибка обработки диапазона '{range_addr}' для стиля '{style_name}': {apply_range_e}")

            except Exception as e:
                logger.error(f"[СТИЛЬ] Ошибка обработки стиля {i+1}: {e}", exc_info=True)

        logger.info(f"[СТИЛЬ] === Конец экспорта стилей для листа '{ws.title}'. Обработано {applied_styles_count} диапазонов стилей. ===")
        return True

    except Exception as e:
        logger.error(f"[СТИЛЬ] Критическая ошибка при экспорте стилей для листа '{ws.title}': {e}", exc_info=True)
        return False

# --- Функции для экспорта диаграмм ---

def export_sheet_charts(ws: Worksheet, charts_data: List[Dict[str, Any]]) -> bool:
    """Экспортирует диаграммы на лист."""
    try:
        logger.info(f"[ДИАГРАММА] === Начало экспорта диаграмм для листа '{ws.title}' ===")
        logger.debug(f"[ДИАГРАММА] Получено {len(charts_data)} диаграмм.")

        if not charts_data:
            logger.info("[ДИАГРАММА] Нет диаграмм для экспорта.")
            return True

        for i, chart_info in enumerate(charts_data):
            logger.debug(f"[ДИАГРАММА] Обработка диаграммы {i+1}: {chart_info}")
            try:
                # 1. Создание объекта диаграммы
                chart_type_id = chart_info.get("type", 1) # 1 - ColumnChart, 2 - LineChart, 3 - BarChart, 5 - PieChart и т.д.
                chart_type_map = {1: BarChart, 2: LineChart, 3: BarChart, 5: PieChart} # Уточнить маппинг
                chart_class = chart_type_map.get(chart_type_id, BarChart) # По умолчанию BarChart
                chart = chart_class()
                logger.debug(f"[ДИАГРАММА] Создана диаграмма типа {chart_class.__name__}")

                # 2. Настройка свойств диаграммы
                if chart_info.get("title"):
                    chart.title = chart_info["title"]
                # if chart_info.get("style") is not None:
                #     chart.style = int(chart_info["style"])
                
                # Настройка осей (пример для BarChart)
                # if hasattr(chart, 'x_axis') and chart_info.get("x_axis_title"):
                #     chart.x_axis.title = chart_info["x_axis_title"]
                # if hasattr(chart, 'y_axis') and chart_info.get("y_axis_title"):
                #     chart.y_axis.title = chart_info["y_axis_title"]

                # 3. Добавление данных
                # Для этого нужно получить данные из связанных таблиц chart_series, chart_data_sources
                # Это упрощенный пример, логику нужно адаптировать под вашу структуру БД
                series_data = chart_info.get("series_data", []) # Предполагаем, что данные уже подготовлены
                for series_item in series_data:
                    try:
                        values_formula = series_item.get("values_formula", "")
                        categories_formula = series_item.get("categories_formula", "")
                        
                        if values_formula:
                            # Предполагаем, что formula это строка вида "Sheet!$A$1:$B$10"
                            parts = values_formula.split('!', 1)
                            if len(parts) == 2:
                                data_sheet_name, data_range = parts
                                # Убедимся, что имя листа существует
                                if data_sheet_name in ws.parent.sheetnames:
                                    data_sheet = ws.parent[data_sheet_name]
                                else:
                                    logger.warning(f"[ДИАГРАММА] Лист '{data_sheet_name}' для данных диаграммы не найден. Используется текущий лист '{ws.title}'.")
                                    data_sheet = ws
                                try:
                                    data_ref = Reference(data_sheet, range_string=data_range)
                                    chart.add_data(data_ref, from_rows=False) # from_rows зависит от структуры данных
                                    logger.debug(f"[ДИАГРАММА] Данные добавлены в диаграмму из {values_formula}")
                                except Exception as ref_e:
                                    logger.error(f"[ДИАГРАММА] Ошибка создания Reference для данных '{values_formula}': {ref_e}")
                        # else:
                        #     logger.debug("[ДИАГРАММА] Нет формулы значений для серии.")

                        # if categories_formula:
                        #     cat_parts = categories_formula.split('!', 1)
                        #     if len(cat_parts) == 2:
                        #         cat_sheet_name, cat_range = cat_parts
                        #         if cat_sheet_name in ws.parent.sheetnames:
                        #             cat_sheet = ws.parent[cat_sheet_name]
                        #         else:
                        #              logger.warning(f"[ДИАГРАММА] Лист '{cat_sheet_name}' для категорий диаграммы не найден. Используется текущий лист '{ws.title}'.")
                        #              cat_sheet = ws
                        #         try:
                        #             cats_ref = Reference(cat_sheet, range_string=cat_range)
                        #             chart.set_categories(cats_ref)
                        #             logger.debug(f"[ДИАГРАММА] Категории добавлены в диаграмму из {categories_formula}")
                        #         except Exception as cat_ref_e:
                        #             logger.error(f"[ДИАГРАММА] Ошибка создания Reference для категорий '{categories_formula}': {cat_ref_e}")
                        # else:
                        #      logger.debug("[ДИАГРАММА] Нет формулы категорий для серии.")
                    
                    except Exception as series_e:
                        logger.error(f"[ДИАГРАММА] Ошибка добавления серии данных в диаграмму {i+1}: {series_e}")

                # 4. Добавление диаграммы на лист
                anchor_cell = chart_info.get("top_left_cell", "A1") # Ячейка для привязки
                try:
                    ws.add_chart(chart, anchor_cell)
                    logger.debug(f"[ДИАГРАММА] Диаграмма {i+1} добавлена в ячейку {anchor_cell}")
                except Exception as add_chart_e:
                    logger.error(f"[ДИАГРАММА] Ошибка добавления диаграммы в ячейку {anchor_cell}: {add_chart_e}")

            except Exception as chart_e:
                logger.error(f"[ДИАГРАММА] Ошибка обработки диаграммы {i+1}: {chart_e}", exc_info=True)
                # Продолжаем с другими диаграммами

        logger.info(f"[ДИАГРАММА] === Конец экспорта диаграмм для листа '{ws.title}' ===")
        return True

    except Exception as e:
        logger.error(f"[ДИАГРАММА] Критическая ошибка при экспорте диаграмм для листа '{ws.title}': {e}", exc_info=True)
        return False

# --- Основная логика экспорта ---

def export_project_to_excel_standalone(project_db_path: Union[str, Path], output_path: Union[str, Path]) -> bool:
    """
    Экспортирует проект из SQLite БД в файл Excel (.xlsx) с использованием openpyxl.
    Объединяет экспорт данных, стилей и диаграмм в один проход.
    
    Args:
        project_db_path (Union[str, Path]): Путь к файлу project_data.db.
        output_path (Union[str, Path]): Путь для сохранения выходного .xlsx файла.
        
    Returns:
        bool: True если экспорт успешен, False в противном случае.
    """
    logger.info("=== НАЧАЛО САМОСТОЯТЕЛЬНОГО ЭКСПОРТА ПРОЕКТА ===")
    logger.info(f"Путь к БД проекта: {project_db_path}")
    logger.info(f"Путь к выходному файлу: {output_path}")

    try:
        project_db_path = Path(project_db_path)
        output_path = Path(output_path)

        if not project_db_path.exists():
            logger.error(f"Файл БД проекта не найден: {project_db_path}")
            return False

        # Импортируем ProjectDBStorage. Путь уже должен быть добавлен.
        try:
            from src.storage.database import ProjectDBStorage
            logger.debug("[ЭКСПОРТ] ProjectDBStorage успешно импортирован.")
        except ImportError as e:
            logger.critical(f"[ЭКСПОРТ] Не удалось импортировать ProjectDBStorage даже после попытки добавить корень проекта в sys.path: {e}")
            return False

        # 1. Подключение к БД
        logger.info("[ЭКСПОРТ] Подключение к БД проекта...")
        storage = ProjectDBStorage(str(project_db_path))
        storage.connect()
        logger.info("[ЭКСПОРТ] Подключение к БД установлено.")

        # 2. Извлечение данных проекта
        logger.info("[ЭКСПОРТ] Извлечение данных проекта из БД...")
        project_data = storage.get_all_data()
        if not project_data:
            logger.error("[ЭКСПОРТ] Не удалось получить данные проекта из БД.")
            storage.disconnect()
            return False
        logger.debug(f"[ЭКСПОРТ] Получены данные проекта: {list(project_data.keys())}")

        # Адаптация к структуре данных ProjectDBStorage
        # Ожидаем project_data = {'sheets': {'SheetName1': {...}, 'SheetName2': {...}}}
        sheets_data = project_data.get("sheets", {})
        if not sheets_data:
            logger.warning("[ЭКСПОРТ] В проекте не найдено листов для экспорта.")
            storage.disconnect()
            # Создаем пустой файл
            wb_empty = Workbook()
            if "Sheet" in wb_empty.sheetnames:
                wb_empty.remove(wb_empty["Sheet"])
            # Создаем один пустой лист, чтобы файл был валидным
            wb_empty.create_sheet(title="Sheet1")
            wb_empty.save(output_path)
            logger.info("[ЭКСПОРТ] Создан пустой файл Excel с одним листом.")
            return True

        # 3. Создание новой книги Excel
        logger.info("[ЭКСПОРТ] Создание новой книги Excel...")
        wb = Workbook()
        # Удаляем дефолтный лист, если он есть
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        logger.info("[ЭКСПОРТ] Новая книга Excel создана.")

        # --- Этап 1: Данные и формулы ---
        logger.info("[ЭКСПОРТ] === Этап 1: Создание структуры и заполнение данными/формулами ===")
        export1_success = True
        # Перебираем листы по имени и данным
        for sheet_name, sheet_info in sheets_data.items():
            logger.info(f"[ЭКСПОРТ] Обработка листа: {sheet_name}")
            # ID листа находится внутри sheet_info
            sheet_id = sheet_info.get("id") # Теперь это должно работать
            if not sheet_id:
                logger.warning(f"[ЭКСПОРТ] ID листа не найден для '{sheet_name}'. Пропущен.")
                continue

            # Создаем лист в новой книге
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(title=sheet_name)
            else:
                ws = wb[sheet_name]

            # Используем уже извлеченные данные из project_data (sheets_data)
            sheet_full_data = sheet_info # sheet_info уже содержит все данные для этого листа

            if not export_sheet_data_and_formulas(ws, sheet_full_data):
                logger.error(f"[ЭКСПОРТ] Ошибка при экспорте данных/формул для листа '{sheet_name}'.")
                export1_success = False # Отмечаем частичный сбой, но продолжаем

        if not export1_success:
             logger.warning("[ЭКСПОРТ] Этап 1 (данные/формул) завершен с ошибками.")

        # --- Этап 2: Стили ---
        logger.info("[ЭКСПОРТ] === Этап 2: Применение стилей ===")
        export2_success = True
        for sheet_name, sheet_info in sheets_data.items():
            logger.info(f"[ЭКСПОРТ] Применение стилей для листа: {sheet_name}")
            sheet_id = sheet_info.get("id") # Получаем ID из данных листа
            if not sheet_id:
                logger.warning(f"[ЭКСПОРТ] ID листа не найден для '{sheet_name}' (стили). Пропущен.")
                continue

            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                styled_ranges_data = sheet_info.get("styled_ranges_data", [])
                if not export_sheet_styles(wb, ws, styled_ranges_data):
                    logger.error(f"[ЭКСПОРТ] Ошибка при экспорте стилей для листа '{sheet_name}'.")
                    export2_success = False
            else:
                 logger.warning(f"[ЭКСПОРТ] Лист '{sheet_name}' не найден в книге для применения стилей.")

        if not export2_success:
             logger.warning("[ЭКСПОРТ] Этап 2 (стили) завершен с ошибками.")

        # --- Этап 3: Диаграммы ---
        logger.info("[ЭКСПОРТ] === Этап 3: Добавление диаграмм ===")
        export3_success = True
        for sheet_name, sheet_info in sheets_data.items():
            logger.info(f"[ЭКСПОРТ] Добавление диаграмм для листа: {sheet_name}")
            sheet_id = sheet_info.get("id") # Получаем ID из данных листа
            if not sheet_id:
                logger.warning(f"[ЭКСПОРТ] ID листа не найден для '{sheet_name}' (диаграммы). Пропущен.")
                continue

            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                charts_data = sheet_info.get("charts_data", [])
                if not export_sheet_charts(ws, charts_data):
                    logger.error(f"[ЭКСПОРТ] Ошибка при экспорте диаграмм для листа '{sheet_name}'.")
                    export3_success = False
            else:
                 logger.warning(f"[ЭКСПОРТ] Лист '{sheet_name}' не найден в книге для добавления диаграмм.")

        if not export3_success:
             logger.warning("[ЭКСПОРТ] Этап 3 (диаграммы) завершен с ошибками.")

        # 4. Сохранение файла
        logger.info(f"[ЭКСПОРТ] === Сохранение финального файла в {output_path} ===")
        try:
            wb.save(output_path)
            logger.info(f"[ЭКСПОРТ] === Файл успешно сохранен: {output_path} ===")
        except Exception as save_e:
            logger.critical(f"[ЭКСПОРТ] Критическая ошибка при сохранении файла: {save_e}", exc_info=True)
            storage.disconnect()
            return False

        # 5. Закрытие соединения с БД
        storage.disconnect()
        logger.info("[ЭКСПОРТ] Соединение с БД закрыто.")

        overall_success = export1_success and export2_success and export3_success
        if overall_success:
            logger.info("=== САМОСТОЯТЕЛЬНЫЙ ЭКСПОРТ ПРОЕКТА ЗАВЕРШЕН УСПЕШНО ===")
        else:
            logger.warning("=== САМОСТОЯТЕЛЬНЫЙ ЭКСПОРТ ПРОЕКТА ЗАВЕРШЕН С ОШИБКАМИ В ОТДЕЛЬНЫХ ЭТАПАХ ===")
        return overall_success

    except Exception as e:
        logger.critical(f"[ЭКСПОРТ] Критическая ошибка при самостоятельном экспорте проекта: {e}", exc_info=True)
        # Пытаемся закрыть соединение, если оно было открыто
        try:
            if 'storage' in locals():
                storage.disconnect()
        except:
            pass
        return False

# --- Точка входа для прямого запуска скрипта ---
if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Самостоятельный экспорт проекта Excel Micro DB.")
    parser.add_argument("project_db_path", help="Путь к файлу project_data.db")
    parser.add_argument("output_path", help="Путь для сохранения выходного .xlsx файла")

    args = parser.parse_args()

    success = export_project_to_excel_standalone(args.project_db_path, args.output_path)

    if success:
        print(f"Экспорт успешно завершен. Файл сохранен в: {args.output_path}")
        sys.exit(0)
    else:
        print(f"Экспорт завершился с ошибкой.")
        sys.exit(1)
