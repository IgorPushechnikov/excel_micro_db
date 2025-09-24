# src/exporter/direct_db_exporter.py
"""
Экспорт проекта Excel Micro DB напрямую из SQLite БД в файл Excel (.xlsx).
Объединяет экспорт данных, стилей и диаграмм в один проход с помощью openpyxl.
"""

import sys
from pathlib import Path
import sqlite3
import json
import logging
from typing import Dict, Any, List, Optional, Tuple, Union

# Импорты openpyxl
from openpyxl import Workbook
from openpyxl.workbook.workbook import Workbook as OpenpyxlWorkbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.cell_range import CellRange
# Импорты для стилей
from openpyxl.styles import (
    Font, Fill, Border, PatternFill, Side, Alignment, Protection, NamedStyle, Color
)
# Импорты для диаграмм - исключены
# from openpyxl.chart import BarChart, LineChart, PieChart, Reference

# --- Настройка базового логирования ---
# Этот базовый логгер будет перенастроен позже в main()
logger = logging.getLogger("direct_db_exporter")


# --- Вспомогательные функции для работы с путями ---
def _add_project_root_to_path_if_needed():
    """Добавляет корень проекта в sys.path, если модуль src не найден."""
    try:
        import src
        logger.debug("[ПУТЬ] Модуль 'src' уже доступен в sys.path.")
    except ImportError:
        logger.debug("[ПУТЬ] Модуль 'src' не найден. Попытка добавить корень проекта в sys.path.")
        current_script_path = Path(__file__).resolve()
        potential_root = current_script_path.parent
        max_levels_up = 10
        # === ИСПРАВЛЕНО: Инициализация project_root_candidate до блока if ===
        # Решает ошибку Pylance: "Элемент "project_root_candidate", возможно, не привязан"
        project_root_candidate: Optional[Path] = None # <--- Инициализация
        root_found = False
        for _ in range(max_levels_up):
            if (potential_root / 'src').is_dir():
                project_root_candidate = potential_root # <--- Присвоение
                root_found = True
                break
            potential_root = potential_root.parent

        if root_found and project_root_candidate:
            project_root_str = str(project_root_candidate)
            if project_root_str not in sys.path:
                sys.path.insert(0, project_root_str)
                logger.debug(f"[ПУТЬ] Корень проекта '{project_root_str}' добавлен в sys.path.")
            else:
                logger.debug(f"[ПУТЬ] Корень проекта '{project_root_str}' уже в sys.path.")
        else:
            logger.warning("[ПУТЬ] Не удалось автоматически определить корень проекта.")


_add_project_root_to_path_if_needed()
# --- Конец вспомогательных функций для путей ---


# --- Функции для работы со стилями ---
def _create_openpyxl_font_from_db_row(font_row: sqlite3.Row) -> Optional[Font]:
    """Создает объект Font openpyxl из строки БД."""
    if not font_row:
        return None
    try:
        font_kwargs = {}
        if 'name' in font_row.keys() and font_row['name']: font_kwargs['name'] = font_row['name']
        if 'sz' in font_row.keys() and font_row['sz'] is not None: font_kwargs['sz'] = float(font_row['sz'])
        font_kwargs['bold'] = bool(font_row['b']) if 'b' in font_row.keys() and font_row['b'] is not None else False
        font_kwargs['italic'] = bool(font_row['i']) if 'i' in font_row.keys() and font_row['i'] is not None else False
        if 'u' in font_row.keys() and font_row['u']: font_kwargs['underline'] = font_row['u']
        font_kwargs['strike'] = bool(font_row['strike']) if 'strike' in font_row.keys() and font_row['strike'] is not None else False
        if 'vert_align' in font_row.keys() and font_row['vert_align']: font_kwargs['vertAlign'] = font_row['vert_align']
        if 'scheme' in font_row.keys() and font_row['scheme']: font_kwargs['scheme'] = font_row['scheme']

        color_kwargs = {}
        if 'color' in font_row.keys() and font_row['color']:
            color_kwargs['rgb'] = font_row['color']
        elif 'color_theme' in font_row.keys() and font_row['color_theme'] is not None:
            color_kwargs['theme'] = int(font_row['color_theme'])
            if 'color_tint' in font_row.keys() and font_row['color_tint'] is not None:
                color_kwargs['tint'] = float(font_row['color_tint'])
        if color_kwargs:
            try:
                font_kwargs['color'] = Color(**color_kwargs)
            except Exception as e:
                logger.warning(f"[СТИЛЬ] Ошибка создания цвета шрифта из {color_kwargs}: {e}")
        logger.debug(f"[СТИЛЬ] Создан Font с параметрами: {font_kwargs}")
        return Font(**font_kwargs)
    except Exception as e:
        logger.error(f"[СТИЛЬ] Ошибка создания Font из строки БД: {e}", exc_info=True)
        return None


def _create_openpyxl_fill_from_db_row(fill_row: sqlite3.Row) -> Optional[Fill]:
    """Создает объект Fill openpyxl из строки БД."""
    if not fill_row or not ('pattern_type' in fill_row.keys() and fill_row['pattern_type']):
        return None
    try:
        fill_kwargs = {'patternType': fill_row['pattern_type']}
        fg_color_kwargs = {}
        if 'fg_color' in fill_row.keys() and fill_row['fg_color']:
            fg_color_kwargs['rgb'] = fill_row['fg_color']
        elif 'fg_color_theme' in fill_row.keys() and fill_row['fg_color_theme'] is not None:
            fg_color_kwargs['theme'] = int(fill_row['fg_color_theme'])
            if 'fg_color_tint' in fill_row.keys() and fill_row['fg_color_tint'] is not None:
                fg_color_kwargs['tint'] = float(fill_row['fg_color_tint'])
        if fg_color_kwargs:
            try:
                fill_kwargs['fgColor'] = Color(**fg_color_kwargs)
            except Exception as e:
                logger.warning(f"[СТИЛЬ] Ошибка создания fgColor из {fg_color_kwargs}: {e}")

        bg_color_kwargs = {}
        if 'bg_color' in fill_row.keys() and fill_row['bg_color']:
            bg_color_kwargs['rgb'] = fill_row['bg_color']
        elif 'bg_color_theme' in fill_row.keys() and fill_row['bg_color_theme'] is not None:
            bg_color_kwargs['theme'] = int(fill_row['bg_color_theme'])
            if 'bg_color_tint' in fill_row.keys() and fill_row['bg_color_tint'] is not None:
                bg_color_kwargs['tint'] = float(fill_row['bg_color_tint'])
        if bg_color_kwargs:
            try:
                fill_kwargs['bgColor'] = Color(**bg_color_kwargs)
            except Exception as e:
                logger.warning(f"[СТИЛЬ] Ошибка создания bgColor из {bg_color_kwargs}: {e}")

        logger.debug(f"[СТИЛЬ] Создан Fill с параметрами: {fill_kwargs}")
        return PatternFill(**fill_kwargs)
    except Exception as e:
        logger.error(f"[СТИЛЬ] Ошибка создания Fill из строки БД: {e}", exc_info=True)
        return None


def _create_openpyxl_side_from_attrs(style: Optional[str], color: Optional[str]) -> Optional[Side]:
    """Создает объект Side openpyxl."""
    if not style and not color: return None
    try:
        side_kwargs = {}
        if style: side_kwargs['style'] = style
        if color:
            try:
                side_kwargs['color'] = Color(rgb=color)
            except Exception as e:
                logger.warning(f"[СТИЛЬ] Ошибка создания цвета Side '{color}': {e}")
        logger.debug(f"[СТИЛЬ] Создан Side с параметрами: {side_kwargs}")
        return Side(**side_kwargs)
    except Exception as e:
        logger.error(f"[СТИЛЬ] Ошибка создания Side из style='{style}', color='{color}': {e}", exc_info=True)
        return None


def _create_openpyxl_border_from_db_row(border_row: sqlite3.Row) -> Border:
    """Создает объект Border openpyxl из строки БД."""
    try:
        sides = {}
        for side_name in ['left', 'right', 'top', 'bottom']:
            style_key = f'{side_name}_style'
            color_key = f'{side_name}_color'
            # ИСПРАВЛЕНО: Правильный синтаксис для sqlite3.Row
            style = border_row[style_key] if style_key in border_row.keys() else None
            color = border_row[color_key] if color_key in border_row.keys() else None
            side_obj = _create_openpyxl_side_from_attrs(style, color)
            if side_obj:
                sides[side_name] = side_obj
                logger.debug(f"[СТИЛЬ] Создана сторона Border {side_name}: {side_obj}")

        # Диагональ
        # ИСПРАВЛЕНО: Правильный синтаксис для sqlite3.Row
        diagonal_style = border_row['diagonal_style'] if 'diagonal_style' in border_row.keys() else None
        diagonal_color = border_row['diagonal_color'] if 'diagonal_color' in border_row.keys() else None
        diagonal_side = _create_openpyxl_side_from_attrs(diagonal_style, diagonal_color)
        if diagonal_side:
            sides['diagonal'] = diagonal_side
            logger.debug(f"[СТИЛЬ] Создана диагональ Border: {diagonal_side}")

        # Другие атрибуты
        # ИСПРАВЛЕНО: Правильный синтаксис для sqlite3.Row
        kwargs = {
            'diagonalUp': bool(border_row['diagonal_up']) if 'diagonal_up' in border_row.keys() and border_row['diagonal_up'] is not None else None,
            'diagonalDown': bool(border_row['diagonal_down']) if 'diagonal_down' in border_row.keys() and border_row['diagonal_down'] is not None else None,
            'outline': bool(border_row['outline']) if 'outline' in border_row.keys() and border_row['outline'] is not None else None
        }
        kwargs.update(sides)
        # Фильтруем None
        filtered_kwargs = {k: v for k, v in kwargs.items() if v is not None}
        logger.debug(f"[СТИЛЬ] Создан Border с параметрами: {filtered_kwargs}")
        # === ДОБАВЛЕНО: # type: ignore[arg-type] для подавления ошибки Pylance ===
        # Решает ошибку Pylance: "Аргумент типа "bool" нельзя присвоить параметру "..." типа "Side | None""
        # если Pylance всё ещё не понимает, что filtered_kwargs содержит правильные типы.
        return Border(**filtered_kwargs) # type: ignore[arg-type]
    except Exception as e:
        logger.error(f"[СТИЛЬ] Ошибка создания Border из строки БД: {e}", exc_info=True)
        return Border()


def _create_openpyxl_alignment_from_db_row(align_row: sqlite3.Row) -> Alignment:
    """Создает объект Alignment openpyxl из строки БД."""
    try:
        align_kwargs = {}
        if 'horizontal' in align_row.keys() and align_row['horizontal']:
            align_kwargs['horizontal'] = align_row['horizontal']
        if 'vertical' in align_row.keys() and align_row['vertical']:
            align_kwargs['vertical'] = align_row['vertical']
        if 'text_rotation' in align_row.keys() and align_row['text_rotation'] is not None:
            align_kwargs['textRotation'] = int(align_row['text_rotation'])
        align_kwargs['wrapText'] = bool(align_row['wrap_text']) if 'wrap_text' in align_row.keys() and align_row[
            'wrap_text'] is not None else False
        align_kwargs['shrinkToFit'] = bool(align_row['shrink_to_fit']) if 'shrink_to_fit' in align_row.keys() and \
                                                                           align_row['shrink_to_fit'] is not None else False
        if 'indent' in align_row.keys() and align_row['indent'] is not None:
            align_kwargs['indent'] = int(align_row['indent'])
        logger.debug(f"[СТИЛЬ] Создан Alignment с параметрами: {align_kwargs}")
        return Alignment(**align_kwargs)
    except Exception as e:
        logger.error(f"[СТИЛЬ] Ошибка создания Alignment из строки БД: {e}", exc_info=True)
        return Alignment()


def _create_openpyxl_protection_from_db_row(prot_row: sqlite3.Row) -> Protection:
    """Создает объект Protection openpyxl из строки БД."""
    try:
        prot_kwargs = {
            'locked': bool(prot_row['locked']) if 'locked' in prot_row.keys() and prot_row['locked'] is not None else True,
            'hidden': bool(prot_row['hidden']) if 'hidden' in prot_row.keys() and prot_row['hidden'] is not None else False
        }
        logger.debug(f"[СТИЛЬ] Создан Protection с параметрами: {prot_kwargs}")
        return Protection(**prot_kwargs)
    except Exception as e:
        logger.error(f"[СТИЛЬ] Ошибка создания Protection из строки БД: {e}", exc_info=True)
        return Protection()


def _fetch_and_create_named_style(conn: sqlite3.Connection, style_id: int,
                                  style_cache: Dict[int, NamedStyle]) -> Optional[NamedStyle]:
    """Получает и создает именованный стиль из БД по его ID, используя кэш."""
    if style_id in style_cache:
        logger.debug(f"[СТИЛЬ] Стиль ID {style_id} найден в кэше.")
        return style_cache[style_id]

    try:
        logger.debug(f"[СТИЛЬ] Запрос стиля ID {style_id} из cell_styles...")
        # 1. Получаем запись из cell_styles
        cursor = conn.execute("SELECT * FROM cell_styles WHERE id = ?", (style_id,))
        style_row = cursor.fetchone()
        if not style_row:
            logger.warning(f"[СТИЛЬ] Стиль с ID {style_id} не найден в cell_styles.")
            return None

        # 2. Генерируем уникальное имя для стиля
        style_name = f"StyleFromDB_{style_id}_{abs(hash(str(dict(style_row)))) % 1000000}"

        # 3. Создаем компоненты стиля
        font_obj = None
        if style_row['font_id'] is not None:
            logger.debug(f"[СТИЛЬ] Запрос шрифта ID {style_row['font_id']}...")
            font_cursor = conn.execute("SELECT * FROM fonts WHERE id = ?", (style_row['font_id'],))
            font_row_data = font_cursor.fetchone()
            if font_row_data: # ИСПРАВЛЕНО: правильная проверка
                try: # Добавлено: отдельный блок try для создания объекта шрифта
                    font_obj = _create_openpyxl_font_from_db_row(font_row_data)
                    logger.debug(f"[СТИЛЬ] Шрифт ID {style_row['font_id']} создан.")
                except Exception as font_create_error: # Добавлено: обработка ошибок создания
                    logger.error(f"[СТИЛЬ] Ошибка создания шрифта ID {style_row['font_id']}: {font_create_error}", exc_info=True)
            else:
                logger.warning(f"[СТИЛЬ] Шрифт ID {style_row['font_id']} не найден.")

        fill_obj = None
        if style_row['fill_id'] is not None:
            logger.debug(f"[СТИЛЬ] Запрос заливки ID {style_row['fill_id']}...")
            fill_cursor = conn.execute("SELECT * FROM pattern_fills WHERE id = ?", (style_row['fill_id'],))
            fill_row_data = fill_cursor.fetchone()
            if fill_row_data: # ИСПРАВЛЕНО
                try: # Добавлено
                    fill_obj = _create_openpyxl_fill_from_db_row(fill_row_data)
                    logger.debug(f"[СТИЛЬ] Заливка ID {style_row['fill_id']} создана.")
                except Exception as fill_create_error: # Добавлено
                    logger.error(f"[СТИЛЬ] Ошибка создания заливки ID {style_row['fill_id']}: {fill_create_error}", exc_info=True)
            else:
                logger.warning(f"[СТИЛЬ] Заливка ID {style_row['fill_id']} не найдена.")

        border_obj = None
        if style_row['border_id'] is not None:
            logger.debug(f"[СТИЛЬ] Запрос границ ID {style_row['border_id']}...")
            border_cursor = conn.execute("SELECT * FROM borders WHERE id = ?", (style_row['border_id'],))
            border_row_data = border_cursor.fetchone()
            if border_row_data: # ИСПРАВЛЕНО
                try: # Добавлено
                    border_obj = _create_openpyxl_border_from_db_row(border_row_data)
                    logger.debug(f"[СТИЛЬ] Границы ID {style_row['border_id']} созданы.")
                except Exception as border_create_error: # Добавлено
                    logger.error(f"[СТИЛЬ] Ошибка создания границ ID {style_row['border_id']}: {border_create_error}", exc_info=True)
            else:
                logger.warning(f"[СТИЛЬ] Границы ID {style_row['border_id']} не найдены.")

        align_obj = None
        if style_row['alignment_id'] is not None:
            logger.debug(f"[СТИЛЬ] Запрос выравнивания ID {style_row['alignment_id']}...")
            align_cursor = conn.execute("SELECT * FROM alignments WHERE id = ?", (style_row['alignment_id'],))
            align_row_data = align_cursor.fetchone()
            if align_row_data: # ИСПРАВЛЕНО
                try: # Добавлено
                    align_obj = _create_openpyxl_alignment_from_db_row(align_row_data)
                    logger.debug(f"[СТИЛЬ] Выравнивание ID {style_row['alignment_id']} создано.")
                except Exception as align_create_error: # Добавлено
                    logger.error(f"[СТИЛЬ] Ошибка создания выравнивания ID {style_row['alignment_id']}: {align_create_error}", exc_info=True)
            else:
                logger.warning(f"[СТИЛЬ] Выравнивание ID {style_row['alignment_id']} не найдено.")

        prot_obj = None
        if style_row['protection_id'] is not None:
            logger.debug(f"[СТИЛЬ] Запрос защиты ID {style_row['protection_id']}...")
            prot_cursor = conn.execute("SELECT * FROM protections WHERE id = ?", (style_row['protection_id'],))
            prot_row_data = prot_cursor.fetchone()
            if prot_row_data: # ИСПРАВЛЕНО
                try: # Добавлено
                    prot_obj = _create_openpyxl_protection_from_db_row(prot_row_data)
                    logger.debug(f"[СТИЛЬ] Защита ID {style_row['protection_id']} создана.")
                except Exception as prot_create_error: # Добавлено
                    logger.error(f"[СТИЛЬ] Ошибка создания защиты ID {style_row['protection_id']}: {prot_create_error}", exc_info=True)
            else:
                logger.warning(f"[СТИЛЬ] Защита ID {style_row['protection_id']} не найдена.")

        # 4. Создаем и кэшируем NamedStyle
        logger.debug(f"[СТИЛЬ] Создание NamedStyle '{style_name}'...")
        try: # Обернем создание NamedStyle в try
            named_style = NamedStyle(name=style_name)
            if font_obj: named_style.font = font_obj
            if fill_obj: named_style.fill = fill_obj
            if border_obj: named_style.border = border_obj
            if align_obj: named_style.alignment = align_obj
            if prot_obj: named_style.protection = prot_obj

            style_cache[style_id] = named_style
            logger.info(f"[СТИЛЬ] Создан и закэширован стиль '{style_name}' (ID: {style_id})")
            return named_style
        except Exception as style_create_error:
            logger.error(f"[СТИЛЬ] Ошибка создания NamedStyle '{style_name}': {style_create_error}", exc_info=True)
            return None # Возвращаем None в случае ошибки создания стиля

    except Exception as e:
        logger.error(f"[СТИЛЬ] Критическая ошибка при получении/создании стиля ID {style_id}: {e}", exc_info=True)
        return None


# --- Функции для экспорта данных ---
def export_sheet_data(conn: sqlite3.Connection, wb: OpenpyxlWorkbook, sheet_id: int, sheet_name: str) -> bool:
    """Экспортирует данные и формулы на лист."""
    try:
        logger.info(f"[ДАННЫЕ] === Начало экспорта данных и формул для листа '{sheet_name}' (ID: {sheet_id}) ===")

        # 1. Создаем или получаем лист
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(title=sheet_name)
            logger.debug(f"[ДАННЫЕ] Создан новый лист '{sheet_name}'.")
        else:
            ws = wb[sheet_name]
            logger.debug(f"[ДАННЫЕ] Лист '{sheet_name}' уже существует.")

        # --- ЭКСПОРТ СЫРЫХ ДАННЫХ ---
        logger.debug("[ДАННЫЕ] --- Экспорт сырых данных ---")
        # 2. Получаем имя таблицы с сырыми данными
        logger.debug("[ДАННЫЕ] Запрос имени таблицы сырых данных...")
        registry_cursor = conn.execute(
            "SELECT table_name FROM raw_data_tables_registry WHERE sheet_id = ?", (sheet_id,)
        )
        registry_row = registry_cursor.fetchone()
        if not registry_row:
            logger.warning(
                f"[ДАННЫЕ] Таблица сырых данных для листа '{sheet_name}' (ID: {sheet_id}) не найдена."
            )
            # Продолжаем, возможно, лист пустой или только с формулами/диаграммами
            raw_data_table_exists = False
            raw_data_table_name = None
        else:
            raw_data_table_exists = True
            raw_data_table_name = registry_row['table_name']
            logger.info(f"[ДАННЫЕ] Имя таблицы сырых данных: {raw_data_table_name}")

        # 3. Заполняем данными, если таблица существует
        if raw_data_table_exists:
            try:
                logger.debug(f"[ДАННЫЕ] Запрос списка столбцов из таблицы {raw_data_table_name}...")
                # Получаем информацию о столбцах
                columns_info_cursor = conn.execute(f"PRAGMA table_info({raw_data_table_name})")
                columns_info = columns_info_cursor.fetchall()
                column_names = [col['name'] for col in columns_info]
                logger.debug(f"[ДАННЫЕ] Столбцы таблицы: {column_names}")

                # Исключаем служебный столбец 'id'
                data_column_names = [name for name in column_names if name.lower() != 'id']
                if not data_column_names:
                    logger.warning(f"[ДАННЫЕ] В таблице {raw_data_table_name} не найдено столбцов с данными (кроме 'id').")
                else:
                    logger.debug(f"[ДАННЫЕ] Столбцы с данными: {data_column_names}")

                    logger.debug(f"[ДАННЫЕ] Запрос данных из таблицы {raw_data_table_name}...")
                    # Формируем SELECT-запрос только для нужных столбцов
                    # Используем двойные кавычки для имен столбцов, чтобы избежать проблем с зарезервированными словами
                    select_columns_clause = ", ".join([f'"{col}"' for col in data_column_names])
                    query = f'SELECT {select_columns_clause} FROM "{raw_data_table_name}"'
                    logger.debug(f"[ДАННЫЕ] SQL-запрос: {query}")
                    data_cursor = conn.execute(query)
                    rows = data_cursor.fetchall()
                    logger.info(f"[ДАННЫЕ] Получено {len(rows)} строк данных.")

                    if rows:
                        # Записываем данные, начиная со строки 2 (строка 1 для заголовков)
                        for row_index_db, row_data in enumerate(rows, start=1): # row_index_db starts at 1
                            row_index_excel = row_index_db + 1 # Excel rows start at 1, data starts at 2
                            logger.debug(f"[ДАННЫЕ] Обработка строки {row_index_db} из БД (Excel строка {row_index_excel})")
                            # row_data - это sqlite3.Row
                            # Используем правильный способ итерации по именам столбцов
                            for col_index_excel, col_name in enumerate(data_column_names, start=1):
                                cell_value = row_data[col_name] # Получаем значение по имени столбца
                                logger.debug(f"[ДАННЫЕ]   -> Запись в ячейку ({row_index_excel}, {col_index_excel}): '{cell_value}' (Тип: {type(cell_value)})")
                                try:
                                    # Непосредственно записываем значение
                                    ws.cell(row=row_index_excel, column=col_index_excel, value=cell_value)
                                except Exception as cell_write_error:
                                    logger.error(f"[ДАННЫЕ] Ошибка записи значения '{cell_value}' в ячейку ({row_index_excel}, {col_index_excel}): {cell_write_error}", exc_info=True)
                    else:
                        logger.debug(f"[ДАННЫЕ] Таблица {raw_data_table_name} пуста.")

            except sqlite3.Error as db_error:
                logger.error(f"[ДАННЫЕ] Ошибка SQLite при работе с таблицей {raw_data_table_name}: {db_error}", exc_info=True)
                # Не возвращаем False, чтобы продолжить экспорт заголовков и формул
            except Exception as data_error:
                logger.error(f"[ДАННЫЕ] Неожиданная ошибка при заполнении данными: {data_error}", exc_info=True)
                # Не возвращаем False, чтобы продолжить экспорт заголовков и формул

        # --- ЭКСПОРТ ЗАГОЛОВКОВ ---
        logger.debug("[ДАННЫЕ] --- Экспорт заголовков ---")
        # 4. Получаем структуру листа для заголовков
        logger.debug("[ДАННЫЕ] Запрос структуры листа для заголовков...")
        sheet_cursor = conn.execute("SELECT structure FROM sheets WHERE sheet_id = ?", (sheet_id,))
        sheet_row = sheet_cursor.fetchone()
        if sheet_row and sheet_row['structure']:
            try:
                structure_data = json.loads(sheet_row['structure'])
                logger.info(f"[ДАННЫЕ] Загружена структура: {len(structure_data)} колонок")
                for col_info in structure_data: # Исправлено: правильное имя переменной
                    col_idx = col_info.get("column_index", 0)
                    if col_idx > 0:  # Excel колонки с 1
                        header = col_info.get("column_name", f"Col{col_idx}")
                        logger.debug(f"[ДАННЫЕ] Запись заголовка '{header}' в ячейку (1, {col_idx})")
                        # Записываем заголовок в первую строку
                        ws.cell(row=1, column=col_idx, value=header)
            except json.JSONDecodeError as json_error:
                logger.error(f"[ДАННЫЕ] Ошибка парсинга структуры листа: {json_error}")
            except Exception as struct_error:
                logger.error(f"[ДАННЫЕ] Ошибка при записи заголовков: {struct_error}", exc_info=True)
        else:
             logger.info(f"[ДАННЫЕ] Структура листа '{sheet_name}' не найдена или пуста.")

        # --- ЭКСПОРТ ФОРМУЛ ---
        logger.debug("[ДАННЫЕ] --- Экспорт формул ---")
        # 5. Добавляем формулы (формулы могут быть в любой строке/столбце)
        logger.debug("[ДАННЫЕ] Запрос формул...")
        formula_cursor = conn.execute("SELECT cell_address AS cell, formula FROM formulas WHERE sheet_id = ?", (sheet_id,))
        formulas = formula_cursor.fetchall()
        logger.info(f"[ДАННЫЕ] Получено {len(formulas)} формул.")
        for formula_row in formulas:
            # Используем правильный синтаксис для sqlite3.Row
            cell_address = formula_row['cell']
            formula_text = formula_row['formula']
            logger.debug(f"[ДАННЫЕ] Запись формулы '{formula_text}' в ячейку {cell_address}")
            try:
                ws[cell_address] = formula_text
            except Exception as formula_error:
                logger.error(f"[ДАННЫЕ] Ошибка записи формулы '{formula_text}' в {cell_address}: {formula_error}", exc_info=True)

        logger.info(f"[ДАННЫЕ] === Завершен экспорт данных и формул для листа '{sheet_name}' ===")
        return True
    except Exception as e:
        logger.error(f"[ДАННЫЕ] Критическая ошибка экспорта данных/формул для листа '{sheet_name}': {e}", exc_info=True)
        return False


# --- Функции для экспорта стилей ---
def export_sheet_styles(conn: sqlite3.Connection, wb: OpenpyxlWorkbook, sheet_id: int, sheet_name: str) -> bool:
    """Экспортирует стили на лист."""
    try:
        logger.info(f"[СТИЛЬ] === Начало экспорта стилей для листа '{sheet_name}' (ID: {sheet_id}) ===")

        if sheet_name not in wb.sheetnames:
            logger.error(f"[СТИЛЬ] Лист '{sheet_name}' не найден в книге.")
            return False
        ws = wb[sheet_name]

        # Кэш для уже созданных NamedStyle
        style_cache: Dict[int, NamedStyle] = {}

        # Получаем все записи стилей для этого листа
        logger.debug("[СТИЛЬ] Запрос записей стилей для листа...")
        styled_ranges_cursor = conn.execute(
            "SELECT style_id, range_address FROM styled_ranges WHERE sheet_id = ?", (sheet_id,)
        )
        styled_ranges = styled_ranges_cursor.fetchall()
        logger.info(f"[СТИЛЬ] Найдено {len(styled_ranges)} записей стилей.")

        if not styled_ranges:
            logger.info(f"[СТИЛЬ] Нет стилей для применения на листе '{sheet_name}'.")
            return True

        applied_count = 0
        for styled_range_row in styled_ranges:
            style_id = styled_range_row['style_id']
            range_address = styled_range_row['range_address']

            logger.debug(f"[СТИЛЬ] Обработка стиля ID {style_id} для диапазона {range_address}")

            # Получаем или создаем стиль
            named_style = _fetch_and_create_named_style(conn, style_id, style_cache)
            if not named_style:
                logger.error(f"[СТИЛЬ] Не удалось получить/создать стиль ID {style_id}. Пропущен.")
                continue

            # Добавляем стиль в книгу, если он еще не добавлен
            # Проверяем по имени, так как wb.named_styles содержит имена
            if named_style.name not in wb.named_styles:
                try:
                    logger.debug(f"[СТИЛЬ] Добавление стиля '{named_style.name}' в книгу...")
                    wb.add_named_style(named_style)
                    logger.debug(f"[СТИЛЬ] Стиль '{named_style.name}' добавлен в книгу.")
                except Exception as e:
                    if "already exists" in str(e).lower():
                        logger.info(
                            f"[СТИЛЬ] Стиль '{named_style.name}' уже существует в книге (ошибка при добавлении).")
                    else:
                        logger.error(f"[СТИЛЬ] Ошибка добавления стиля '{named_style.name}': {e}")
            # else:
            #     logger.debug(f"[СТИЛЬ] Стиль '{named_style.name}' уже в книге.")

            # Применяем стиль к диапазону
            # === ИСПРАВЛЕНО: Упрощенная и надежная логика итерации ===
            # Решает ошибку Pylance: ""MergedCell" не является итерируемым"
            try:
                logger.debug(f"[СТИЛЬ] Применение стиля '{named_style.name}' к диапазону {range_address}...")
                
                # Получаем диапазон. Может быть Cell или tuple[tuple[Cell]]
                cell_range_object = ws[range_address] 
                
                # Создаем плоский список ячеек для стилизации
                cells_to_style = []
                
                # Проверяем, является ли объект итерируемым (tuple, list и т.д.)
                if isinstance(cell_range_object, (list, tuple)):
                    # Это диапазон ячеек, итерируемся по строкам и столбцам
                    logger.debug(f"[СТИЛЬ] Диапазон {range_address} - это коллекция.")
                    for row in cell_range_object:
                        if isinstance(row, (list, tuple)): # Строка диапазона
                            cells_to_style.extend(row)
                        else: # Отдельная ячейка в "плоском" кортеже
                            cells_to_style.append(row)
                else:
                    # Это отдельная ячейка (openpyxl.cell.cell.Cell)
                    logger.debug(f"[СТИЛЬ] Диапазон {range_address} - это одиночная ячейка {cell_range_object.coordinate}.")
                    cells_to_style = [cell_range_object] # Оборачиваем в список

                logger.debug(f"[СТИЛЬ] Всего ячеек для стиля в {range_address}: {len(cells_to_style)}")
                styled_in_range = 0
                for cell in cells_to_style:
                    try:
                        # Применяем стиль по имени - это правильный способ
                        cell.style = named_style.name
                        styled_in_range += 1
                        logger.debug(f"[СТИЛЬ] Стиль '{named_style.name}' применен к {cell.coordinate}")
                    except Exception as e:
                        logger.error(
                            f"[СТИЛЬ] Ошибка применения стиля '{named_style.name}' к {cell.coordinate}: {e}",
                            exc_info=True)

                if styled_in_range > 0:
                    applied_count += 1
                logger.info(
                    f"[СТИЛЬ] Стиль ID {style_id} применен к диапазону {range_address}. Стилизовано ячеек: {styled_in_range}")
            except Exception as e:
                logger.error(f"[СТИЛЬ] Критическая ошибка обработки диапазона {range_address}: {e}", exc_info=True)

        logger.info(
            f"[СТИЛЬ] === Завершен экспорт стилей для листа '{sheet_name}'. Применено к {applied_count}/{len(styled_ranges)} диапазонам. ===")
        return True
    except Exception as e:
        logger.critical(f"[СТИЛЬ] Критическая ошибка экспорта стилей для листа '{sheet_name}': {e}", exc_info=True)
        return False


# --- Функции для экспорта диаграмм ---
# ИСКЛЮЧЕНЫ из текущей версии
def export_sheet_charts(conn: sqlite3.Connection, wb: OpenpyxlWorkbook, sheet_id: int, sheet_name: str) -> bool:
    """Экспортирует диаграммы на лист. (ИСКЛЮЧЕНО)"""
    logger.info(f"[ДИАГРАММА] === Экспорт диаграмм для листа '{sheet_name}' ИСКЛЮЧЕН ===")
    return True # Всегда возвращаем True


# --- Основная логика экспорта ---
def export_project_from_db(db_path: Union[str, Path], output_path: Union[str, Path]) -> bool:
    """
    Экспортирует проект из SQLite БД в файл Excel (.xlsx).
    """
    logger.info("=== НАЧАЛО ЭКСПОРТА ПРОЕКТА ИЗ БД ===")
    logger.info(f"Путь к БД проекта: {db_path}")
    logger.info(f"Путь к выходному файлу: {output_path}")

    db_path = Path(db_path)
    output_path = Path(output_path)

    if not db_path.exists():
        logger.error(f"Файл БД проекта не найден: {db_path}")
        return False

    conn = None
    try:
        # 1. Подключение к БД
        logger.info("[ЭКСПОРТ] Подключение к БД проекта...")
        conn = sqlite3.connect(str(db_path))
        conn.row_factory = sqlite3.Row
        logger.info("[ЭКСПОРТ] Подключение к БД установлено.")

        # 2. Создание новой книги Excel
        logger.info("[ЭКСПОРТ] Создание новой книги Excel...")
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        logger.info("[ЭКСПОРТ] Новая книга Excel создана.")

        # 3. Получение списка листов
        logger.debug("[ЭКСПОРТ] Запрос списка листов...")
        sheet_cursor = conn.execute("SELECT sheet_id AS id, name FROM sheets ORDER BY sheet_id")
        sheets = sheet_cursor.fetchall()
        logger.info(f"[ЭКСПОРТ] Найдено {len(sheets)} листов для экспорта.")

        if not sheets:
            logger.warning("[ЭКСПОРТ] В проекте не найдено листов. Создается пустой файл.")
            wb.create_sheet(title="EmptySheet")
            wb.save(output_path)
            if conn:
                conn.close()
            logger.info(f"[ЭКСПОРТ] Пустой файл сохранен: {output_path}")
            return True

        # --- Этап 1: Данные и формулы ---
        logger.info("[ЭКСПОРТ] === Этап 1: Создание структуры и заполнение данными/формулами ===")
        data_success = True
        for sheet_row in sheets:
            sheet_id = sheet_row['id']
            sheet_name = sheet_row['name']
            logger.info(f"[ЭКСПОРТ] Обработка листа: {sheet_name} (ID: {sheet_id})")
            if not export_sheet_data(conn, wb, sheet_id, sheet_name):
                logger.error(f"[ЭКСПОРТ] Ошибка при экспорте данных/формул для листа '{sheet_name}'.")
                data_success = False  # Отмечаем частичную ошибку, но продолжаем

        # --- Этап 2: Стили ---
        logger.info("[ЭКСПОРТ] === Этап 2: Применение стилей ===")
        style_success = True
        for sheet_row in sheets:
            sheet_id = sheet_row['id']
            sheet_name = sheet_row['name']
            logger.info(f"[ЭКСПОРТ] Применение стилей для листа: {sheet_name} (ID: {sheet_id})")
            if not export_sheet_styles(conn, wb, sheet_id, sheet_name):
                logger.error(f"[ЭКСПОРТ] Ошибка при экспорте стилей для листа '{sheet_name}'.")
                style_success = False

        # --- Этап 3: Диаграммы (ИСКЛЮЧЕН) ---
        logger.info("[ЭКСПОРТ] === Этап 3: Добавление диаграмм ===")
        logger.info("[ЭКСПОРТ] Экспорт диаграмм ИСКЛЮЧЕН.")
        chart_success = True # Считаем успешным, так как исключено

        # 4. Сохранение файла
        logger.info(f"[ЭКСПОРТ] === Сохранение финального файла в {output_path} ===")
        wb.save(output_path)
        logger.info(f"[ЭКСПОРТ] === Файл успешно сохранен: {output_path} ===")

        # 5. Закрытие соединения
        if conn:
            conn.close()
        logger.info("[ЭКСПОРТ] Соединение с БД закрыто.")

        overall_success = data_success and style_success and chart_success
        if overall_success:
            logger.info("=== ЭКСПОРТ ПРОЕКТА ИЗ БД ЗАВЕРШЕН УСПЕШНО ===")
        else:
            logger.warning("=== ЭКСПОРТ ПРОЕКТА ИЗ БД ЗАВЕРШЕН С ОШИБКАМИ НА ОТДЕЛЬНЫХ ЭТАПАХ ===")
        return overall_success

    except Exception as e:
        logger.critical(f"[ЭКСПОРТ] Критическая ошибка при экспорте проекта: {e}", exc_info=True)
        if conn:
            try:
                conn.close()
            except:
                pass
        return False


# --- Точка входа для прямого запуска ---
if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="Экспорт проекта Excel Micro DB напрямую из БД.",
        formatter_class=argparse.RawTextHelpFormatter
    )
    parser.add_argument("db_path", help="Путь к файлу project_data.db")
    parser.add_argument("output_path", help="Путь для сохранения выходного .xlsx файла")
    parser.add_argument("--log-file", help="Путь к файлу для сохранения лога (по умолчанию лог только в консоль)")

    args = parser.parse_args()

    # --- Настройка логирования ---
    log_level = logging.DEBUG
    log_format = '%(asctime)s - %(name)s - %(levelname)s - %(message)s'

    # Создаем логгер для этого модуля
    logger = logging.getLogger("direct_db_exporter")
    logger.setLevel(log_level)

    # Очищаем существующие обработчики (на случай повторного запуска в одном процессе)
    if logger.hasHandlers():
        logger.handlers.clear()

    # Создаем обработчик для вывода в консоль
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(log_level)
    console_formatter = logging.Formatter(log_format)
    console_handler.setFormatter(console_formatter)
    logger.addHandler(console_handler)

    # Создаем обработчик для вывода в файл, если указан --log-file
    if args.log_file:
        try:
            log_file_path = Path(args.log_file)
            log_file_path.parent.mkdir(parents=True, exist_ok=True) # Создаем директории, если нужно
            file_handler = logging.FileHandler(log_file_path, encoding='utf-8')
            file_handler.setLevel(log_level)
            file_formatter = logging.Formatter(log_format)
            file_handler.setFormatter(file_formatter)
            logger.addHandler(file_handler)
            logger.info(f"Логирование настроено. Лог будет записываться в файл: {log_file_path}")
        except Exception as e:
            logger.error(f"Ошибка при настройке логирования в файл {args.log_file}: {e}")
            # Продолжаем работу, используя только консольный лог

    logger.info("=== ЗАПУСК СКРИПТА ЭКСПОРТА ===")
    # --- Конец настройки логирования ---

    success = export_project_from_db(args.db_path, args.output_path)

    if success:
        logger.info(f"Экспорт успешно завершен. Файл сохранен в: {args.output_path}")
        sys.exit(0)
    else:
        logger.error(f"Экспорт завершился с ошибкой.")
        sys.exit(1)

# --- Функция для удобного вызова из CLI/AppController ---
# Исправленная функция export_project
def export_project(project_db_path: str, output_excel_path: str) -> bool:
    """
    Удобная функция для экспорта проекта.
    
    Args:
        project_db_path (str): Путь к файлу БД проекта.
        output_excel_path (str): Путь к выходному .xlsx файлу.
        
    Returns:
        bool: True, если экспорт успешен, иначе False.
    """
    # Используем существующую функцию экспорта
    try:
        success = export_project_from_db(project_db_path, output_excel_path)
        return success
    except Exception as e:
        # Используйте logger, если он импортирован и настроен в этом файле
        # logger.error(f"Ошибка при экспорте проекта: {e}", exc_info=True)
        # Если logger недоступен, можно использовать print (но лучше заменить на logger)
        print(f"Ошибка при экспорте проекта '{project_db_path}' -> '{output_excel_path}': {e}")
        return False

# Пример использования (если файл запускается напрямую)
if __name__ == "__main__":
    import sys
    if len(sys.argv) != 3:
        print("Использование: python direct_db_exporter.py <project_db_path> <output_excel_path>")
        sys.exit(1)
    
    project_db_path = sys.argv[1]
    output_excel_path = sys.argv[2]
    
    if export_project(project_db_path, output_excel_path):
        print(f"Проект успешно экспортирован в {output_excel_path}")
    else:
        print(f"Ошибка при экспорте проекта в {output_excel_path}")
        sys.exit(1)
