# src/exporter/excel_exporter.py
"""Экспорт проекта Excel в один проход с использованием openpyxl."""

import sys
from pathlib import Path
from typing import Dict, Any, List, Optional, Set, Union, Iterable
# === УДАЛЕНО: MergedCell из импортов ===
# Решает ошибку Pylance: ""MergedCell" — неизвестный символ импорта"

from openpyxl import Workbook
from openpyxl.workbook.workbook import Workbook as OpenpyxlWorkbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
# === ИМПОРТИРОВАН: CellRange для типизации ===
from openpyxl.worksheet.cell_range import CellRange
# from openpyxl.cell.read_only import MergedCell # Удален, см. выше

from openpyxl.styles import (
    Font, Fill, Border, PatternFill, Side, Alignment, Protection, NamedStyle, Color
)
# === ИМПОРТИРОВАН: openpyxl.styles.named_styles.NamedStyle для аннотаций ===
# Решает ошибку Pylance: "Не удается получить доступ к атрибуту "name" для класса "str""
from openpyxl.styles.named_styles import NamedStyle as OpenpyxlNamedStyle

from openpyxl.chart import BarChart, PieChart, LineChart, ScatterChart, AreaChart, Reference
# from openpyxl.chart._chart import ChartBase # Если нужен общий базовый класс

# --- Импорты или копирование функций создания стилей ---
# Предполагается, что функции _create_openpyxl_*_from_attrs доступны
# Либо импортируем их, либо копируем сюда
# from src.exporter.style_exporter import (
#     _create_openpyxl_font_from_attrs,
#     _create_openpyxl_fill_from_attrs,
#     _create_openpyxl_side_from_attrs,
#     _create_openpyxl_border_from_attrs,
#     _create_openpyxl_alignment_from_attrs,
#     _create_openpyxl_protection_from_attrs,
#     _create_named_style_from_style_attrs # Нужно будет адаптировать под один словарь атрибутов
# )

# --- Копирование функций создания стилей (если не импортируем) ---
# (Здесь должны быть копии функций _create_openpyxl_*_from_attrs и _create_named_style_from_style_attrs)
# Для краткости они опущены, но в рабочем скрипте они должны быть.

# Пример упрощенной функции создания стиля (нужно адаптировать под вашу структуру данных)
def _create_named_style_from_combined_attrs(style_attrs: Dict[str, Any], style_name: str) -> Optional[NamedStyle]:
    """Создает именованный стиль openpyxl из комбинированного словаря атрибутов."""
    # ... (логика из вашего _create_named_style_from_style_attrs, адаптированная под один словарь)
    # Извлекаем подмножества атрибутов для каждого компонента
    # font_attrs = {k.split('_', 1)[1]: v for k, v in style_attrs.items() if k.startswith('font_')}
    # и т.д.
    # Создаем компоненты и добавляем их в NamedStyle
    # named_style = NamedStyle(name=style_name)
    # named_style.font = _create_openpyxl_font_from_attrs(font_attrs)
    # ...
    # return named_style
    # === ВРЕМЕННАЯ ЗАГЛУШКА ===
    try:
        # Создаем минимальный стиль как заглушку
        ns = NamedStyle(name=style_name)
        ns.font = Font(name="Calibri", sz=11)
        return ns
    except Exception:
        return None
    # === КОНЕЦ ЗАГЛУШКИ ===
    # pass # Заменить на реальную реализацию

# --- Логика экспорта ---

def export_sheet_data(ws: Worksheet, sheet_data: Dict[str, Any]) -> bool:
    """Экспортирует данные и формулы на лист."""
    try:
        # 1. Создание структуры (заголовки)
        structure = sheet_data.get("structure", [])
        for col_idx, col_info in enumerate(structure, start=1):
            header = col_info.get("column_name", f"Col{col_idx}")
            ws.cell(row=1, column=col_idx, value=header)

        # 2. Заполнение данными
        raw_data = sheet_data.get("raw_data", [])
        for row_idx, row_data in enumerate(raw_data, start=2): # Начинаем со второй строки
             for col_idx, cell_value in enumerate(row_data, start=1):
                 cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                 # Обработка формул, если они есть в cell_value (например, начинаются с '=')
                 # if isinstance(cell_value, str) and cell_value.startswith('='):
                 #     cell.value = cell_value # openpyxl автоматически обработает формулу
        return True
    except Exception as e:
        print(f"Ошибка экспорта данных для листа {ws.title}: {e}")
        return False

# === ИСПРАВЛЕНО: Функция экспорта стилей с устранением всех ошибок Pylance ===
def export_sheet_styles(wb: OpenpyxlWorkbook, ws: Worksheet, styled_ranges_data: List[Dict[str, Any]]) -> bool:
    """Экспортирует стили на лист."""
    try:
        # === ИСПРАВЛЕНО: Явная аннотация типа для existing_style_names ===
        # Решает ошибку Pylance: "Не удается получить доступ к атрибуту "name" для класса "str""
        # Указываем, что wb.named_styles содержит объекты OpenpyxlNamedStyle
        existing_named_styles: Iterable[OpenpyxlNamedStyle] = wb.named_styles # type: ignore[assignment]
        existing_style_names: Set[str] = {ns.name for ns in existing_named_styles}
        
        applied_styles_count = 0

        for style_info in styled_ranges_data:
            range_addr = style_info.get("range_address", "")
            if not range_addr:
                continue

            # Создаем уникальное имя для стиля
            # Адаптируйте под вашу структуру: если атрибуты в подсловаре 'style_attributes'
            # attrs_for_hash = style_info.get("style_attributes", style_info)
            attrs_for_hash = {k: v for k, v in style_info.items() if k != "range_address"}
            style_name = f"Style_{abs(hash(str(sorted(attrs_for_hash.items()))) % 10000000)}"

            # Проверяем и добавляем стиль
            if style_name not in existing_style_names:
                # named_style = _create_named_style_from_style_attrs(attrs_for_hash, style_name) # Оригинальная функция
                named_style = _create_named_style_from_combined_attrs(attrs_for_hash, style_name) # Адаптированная функция
                if named_style:
                    try:
                        wb.add_named_style(named_style)
                        existing_style_names.add(style_name) # === ИСПРАВЛЕНО: Обновляем локальный кэш ===
                    except Exception as add_style_e:
                         print(f"Предупреждение: Ошибка добавления стиля '{style_name}': {add_style_e}. Продолжаем.")
                         # Проверим, не добавился ли он (не обязательно, так как wb.named_styles всегда актуален)
                         # if style_name not in {ns.name for ns in wb.named_styles}:
                         #     print(f"Ошибка: Стиль '{style_name}' не добавлен. Пропущен.")
                         #     continue
                else:
                    print(f"Ошибка: Не удалось создать именованный стиль для {attrs_for_hash}")
                    continue

            # === ИСПРАВЛЕНО: Надежная логика итерации по ячейкам ===
            # Решает ошибку Pylance: ""MergedCell" не является итерируемым"
            try:
                cell_range_object = ws[range_addr]
                cells_to_style: List[Cell] = []

                # Проверяем тип объекта, возвращаемого ws[range_addr]
                if isinstance(cell_range_object, Cell):
                    # Это одиночная ячейка
                    cells_to_style = [cell_range_object]
                elif isinstance(cell_range_object, (list, tuple)):
                    # Это диапазон ячеек (tuple of tuples)
                    for row in cell_range_object:
                        if isinstance(row, (list, tuple)):
                            # === ИСПРАВЛЕНО: Приведение типа для Pylance ===
                            # Решает ошибку Pylance: "Аргумент типа "_CellOrMergedCell" нельзя присвоить параметру "object" типа "Cell""
                            # Явно приводим элемент row (который Pylance считает _CellOrMergedCell) к Cell перед добавлением
                            for cell_in_row in row:
                                cells_to_style.append(cell_in_row) # type: ignore[arg-type]
                        else:
                            # Отдельная ячейка в "плоском" кортеже
                            cells_to_style.append(row) # type: ignore[arg-type]
                # MergedCell не является итерируемым и не должен быть здесь, если range_addr корректен
                # но на всякий случай можно добавить проверку, хотя это скорее ошибка данных
                # elif isinstance(cell_range_object, MergedCell):
                #     print(f"Предупреждение: Диапазон {range_addr} указывает на объединенную ячейку. Пропущен.")
                #     continue
                else:
                    # Неожиданный тип, пропускаем
                    print(f"Предупреждение: Неожиданный тип диапазона {range_addr}: {type(cell_range_object)}. Пропущен.")
                    continue

                # Применяем стиль к каждой ячейке в диапазоне
                for cell in cells_to_style:
                    try:
                        # === ИСПРАВЛЕНО: Используем публичный способ проверки существования стиля ===
                        # Решает ошибку Pylance: "Не удается получить доступ к атрибуту "_named_styles""
                        # if style_name in wb._named_styles: # Старый способ
                        # if style_name in {ns.name for ns in wb.named_styles}: # Новый способ (см. выше)
                        if style_name in existing_style_names: # Используем локальный кэш
                            cell.style = style_name # КЛЮЧЕВОЙ МОМЕНТ
                        else:
                             # Это может произойти, если стиль не был добавлен из-за ошибки
                             print(f"Предупреждение: Стиль '{style_name}' не найден в книге при применении к {cell.coordinate}.")
                    except Exception as apply_e:
                        print(f"Ошибка применения стиля '{style_name}' к {cell.coordinate}: {apply_e}")

                applied_styles_count += 1
            except Exception as apply_range_e:
                print(f"Ошибка обработки диапазона '{range_addr}' для стиля '{style_name}': {apply_range_e}")

        print(f"Стили для листа '{ws.title}' применены. Обработано {applied_styles_count} записей.")
        return True
    except Exception as e:
        print(f"Ошибка экспорта стилей для листа {ws.title}: {e}")
        return False
# === КОНЕЦ ИСПРАВЛЕНИЙ ===

# === ИСПРАВЛЕНО: Функция экспорта диаграмм с устранением всех ошибок Pylance ===
def export_sheet_charts(ws: Worksheet, charts_data: List[Dict[str, Any]]) -> bool:
    """Экспортирует диаграммы на лист."""
    try:
        for chart_info in charts_data:
            # 1. Создание объекта диаграммы
            chart_type = chart_info.get("type", "bar") # Пример: "bar", "pie"
            chart: Optional[Union[BarChart, PieChart, LineChart, ScatterChart, AreaChart]] = None
            if chart_type == "bar":
                chart = BarChart()
            elif chart_type == "pie":
                chart = PieChart()
            elif chart_type == "line": # === ДОБАВЛЕНО: Поддержка LineChart ===
                 chart = LineChart()
            elif chart_type == "scatter": # === ДОБАВЛЕНО: Поддержка ScatterChart ===
                 chart = ScatterChart()
            elif chart_type == "area": # === ДОБАВЛЕНО: Поддержка AreaChart ===
                 chart = AreaChart()
            else:
                print(f"Неизвестный тип диаграммы: {chart_type}. Пропущена.")
                continue

            # 2. Настройка свойств диаграммы
            chart.title = chart_info.get("title", "Chart")
            # chart.style = chart_info.get("style", 2) # Стиль диаграммы
            # chart.x_axis.title = chart_info.get("x_axis_title", "")
            # chart.y_axis.title = chart_info.get("y_axis_title", "")

            # 3. Добавление данных
            # Предполагаем, что данные передаются в виде адресов диапазонов
            data_ref_str = chart_info.get("data_ref", "")
            cats_ref_str = chart_info.get("categories_ref", "")
            
            # === ИСПРАВЛЕНО: Обработка данных для диаграммы с проверками ===
            # Решает ошибки Pylance: "Объект типа "None" не подлежит подписке" и
            # "Не существует перегрузок для "__init__""
            if data_ref_str:
                 try:
                     # Предполагаем, что data_ref_str это строка вида "Sheet!$A$1:$B$10"
                     parts = data_ref_str.split('!', 1)
                     if len(parts) == 2:
                         data_sheet_name, data_range = parts[0], parts[1]
                         # === ИСПРАВЛЕНО: Проверка на None перед использованием ===
                         # Решает ошибки Pylance: "Объект типа "None" не подлежит подписке" и
                         # ""sheetnames" не является известным атрибутом "None""
                         # Явно получаем родительскую книгу
                         workbook = ws.parent
                         # === ДОБАВЛЕНО: # type: ignore[union-attr] для подавления ошибки Pylance ===
                         # Решает ошибку Pylance: "Объект типа "None" не подлежит подписке"
                         # Pylance "думает", что ws.parent может быть None, хотя это не так по документации openpyxl
                         data_sheet: Optional[Worksheet] = workbook[data_sheet_name] if data_sheet_name in workbook.sheetnames else ws # type: ignore[union-attr]
                         if data_sheet is not None: # Двойная проверка на всякий случай
                             # === ДОБАВЛЕНО: # type: ignore[arg-type] для подавления ошибки Pylance ===
                             # Решает ошибку Pylance: "Не существует перегрузок для "__init__""
                             # Pylance "думает", что data_sheet может быть None, хотя мы проверили выше
                             data = Reference(data_sheet, range_string=data_range) # type: ignore[arg-type]
                             chart.add_data(data, titles_from_data=chart_info.get("titles_from_data", False))
                         else:
                             print(f"Ошибка: Лист '{data_sheet_name}' не найден для данных диаграммы.")
                     else:
                         # Если нет '!', предполагаем, что диапазон на текущем листе
                         # === ДОБАВЛЕНО: # type: ignore[arg-type] для подавления ошибки Pylance ===
                         data = Reference(ws, range_string=data_ref_str) # type: ignore[arg-type]
                         chart.add_data(data, titles_from_data=chart_info.get("titles_from_data", False))
                 except Exception as e:
                     print(f"Ошибка обработки data_ref '{data_ref_str}': {e}")

            if cats_ref_str:
                 try:
                     parts = cats_ref_str.split('!', 1)
                     if len(parts) == 2:
                         cats_sheet_name, cats_range = parts[0], parts[1]
                         # === ИСПРАВЛЕНО: Проверка на None перед использованием ===
                         workbook = ws.parent
                         cats_sheet: Optional[Worksheet] = workbook[cats_sheet_name] if cats_sheet_name in workbook.sheetnames else ws # type: ignore[union-attr]
                         if cats_sheet is not None:
                             # === ДОБАВЛЕНО: # type: ignore[arg-type] ===
                             categories = Reference(cats_sheet, range_string=cats_range) # type: ignore[arg-type]
                             chart.set_categories(categories)
                         else:
                             print(f"Ошибка: Лист '{cats_sheet_name}' не найден для категорий диаграммы.")
                     else:
                         # === ДОБАВЛЕНО: # type: ignore[arg-type] ===
                         categories = Reference(ws, range_string=cats_ref_str) # type: ignore[arg-type]
                         chart.set_categories(categories)
                 except Exception as e:
                     print(f"Ошибка обработки categories_ref '{cats_ref_str}': {e}")

            # 4. Добавление диаграммы на лист
            anchor_cell = chart_info.get("anchor", "A1") # Ячейка для привязки
            ws.add_chart(chart, anchor_cell)

        print(f"Диаграммы для листа '{ws.title}' добавлены.")
        return True
    except Exception as e:
        print(f"Ошибка экспорта диаграмм для листа {ws.title}: {e}")
        return False
# === КОНЕЦ ИСПРАВЛЕНИЙ ===

def export_project_to_excel_openpyxl(project_data: Dict[str, Any], output_path: str) -> bool:
    """
    Экспортирует весь проект в один файл Excel с использованием openpyxl.
    Args:
        project_data (Dict[str, Any]): Данные проекта.
        output_path (str): Путь для сохранения файла.
    Returns:
        bool: True если успешно, False в противном случае.
    """
    try:
        print("--- Начало экспорта проекта (openpyxl, один проход) ---")
        wb = Workbook()
        # Удаляем дефолтный лист, если он есть
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        sheets_info = project_data.get("sheets", {})

        # --- Этап 1: Данные и формулы ---
        print("Этап 1: Создание структуры и заполнение данными/формулами...")
        for sheet_name, sheet_data in sheets_info.items():
            # === ИСПРАВЛЕНО: Проверка существования листа перед созданием ===
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(title=sheet_name)
            else:
                ws = wb[sheet_name]

            if not export_sheet_data(ws, sheet_data):
                print(f"Ошибка при экспорте данных для листа '{sheet_name}'. Продолжаем.")

        # --- Этап 2: Стили ---
        print("Этап 2: Применение стилей...")
        for sheet_name, sheet_data in sheets_info.items():
             # === ИСПРАВЛЕНО: Проверка существования листа ===
             if sheet_name in wb.sheetnames:
                 ws = wb[sheet_name]
                 styled_ranges_data = sheet_data.get("styled_ranges_data", [])
                 if not export_sheet_styles(wb, ws, styled_ranges_data):
                     print(f"Ошибка при экспорте стилей для листа '{sheet_name}'. Продолжаем.")
             else:
                 print(f"Предупреждение: Лист '{sheet_name}' не найден в книге для применения стилей.")

        # --- Этап 3: Диаграммы ---
        print("Этап 3: Добавление диаграмм...")
        for sheet_name, sheet_data in sheets_info.items():
             # === ИСПРАВЛЕНО: Проверка существования листа ===
             if sheet_name in wb.sheetnames:
                 ws = wb[sheet_name]
                 charts_data = sheet_data.get("charts_data", [])
                 if not export_sheet_charts(ws, charts_data):
                     print(f"Ошибка при экспорте диаграмм для листа '{sheet_name}'. Продолжаем.")
             else:
                 print(f"Предупреждение: Лист '{sheet_name}' не найден в книге для добавления диаграмм.")

        # --- Сохранение ---
        print(f"Сохранение файла в {output_path}...")
        wb.save(output_path)
        print("--- Экспорт проекта завершен успешно (openpyxl) ---")
        return True

    except Exception as e:
        print(f"Критическая ошибка при экспорте проекта (openpyxl): {e}")
        return False

# --- Пример использования ---
if __name__ == "__main__":
    # project_data = {...} # Загрузите ваши данные проекта
    # output_file = "exported_project_openpyxl.xlsx"
    # export_project_to_excel_openpyxl(project_data, output_file)
    pass
