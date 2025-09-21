# src/exporter/style_exporter.py
"""Модуль для экспорта стилей ячеек листа Excel."""
import sys
from pathlib import Path
from typing import Dict, Any, List, Optional, Union, Iterable, Set # Добавлены Set, Iterable
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from openpyxl.cell.cell import Cell # Импортируем Cell напрямую
from openpyxl.worksheet.cell_range import CellRange # Импортируем CellRange напрямую
# Импорты для стилей
from openpyxl.styles import (
    Font, Fill, Border, PatternFill, Side, Alignment, Protection, NamedStyle, Color
)
# === ИСПРАВЛЕНО: Импорт ChartBase ===
# ChartBase определен в openpyxl.chart._chart
from openpyxl.chart._chart import ChartBase
# === КОНЕЦ ИСПРАВЛЕНИЙ ===
# Добавляем корень проекта в путь поиска модулей если нужно
project_root = Path(__file__).parent.parent.parent
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))
from src.utils.logger import get_logger
logger = get_logger(__name__)
# === ИЗМЕНЕНО: Функция для создания Font ===
def _create_openpyxl_font_from_attrs(font_db_attrs: Dict[str, Any]) -> Optional[Font]:
    """
    Создает объект Font openpyxl из атрибутов БД.
    Args:
        font_db_attrs (Dict[str, Any]): Атрибуты шрифта из БД, например,
                                       {'font_name': 'Calibri', 'font_sz': 11.0, 'font_b': 1, ...}.
    Returns:
        Optional[Font]: Созданный объект Font или None в случае ошибки.
    """
    if not font_db_attrs:
        return None
    try:
        font_kwargs = {}
        # Преобразуем атрибуты из БД в аргументы Font
        # Имя шрифта
        if 'font_name' in font_db_attrs and font_db_attrs['font_name'] is not None:
            font_kwargs['name'] = font_db_attrs['font_name']
        # Размер шрифта
        if 'font_sz' in font_db_attrs and font_db_attrs['font_sz'] is not None:
            font_kwargs['sz'] = float(font_db_attrs['font_sz'])
        # Жирность (bold)
        if 'font_b' in font_db_attrs:
            font_kwargs['bold'] = bool(font_db_attrs['font_b'])
        # Курсив (italic)
        if 'font_i' in font_db_attrs:
            font_kwargs['italic'] = bool(font_db_attrs['font_i'])
        # Подчеркивание
        if 'font_u' in font_db_attrs and font_db_attrs['font_u'] is not None:
            font_kwargs['underline'] = font_db_attrs['font_u'] # 'single', 'double' и т.д.
        # Зачеркнутый
        if 'font_strike' in font_db_attrs:
             font_kwargs['strike'] = bool(font_db_attrs['font_strike'])
        # Вертикальное выравнивание текста в строке
        if 'font_vert_align' in font_db_attrs and font_db_attrs['font_vert_align'] is not None:
             font_kwargs['vertAlign'] = font_db_attrs['font_vert_align'] # 'superscript', 'subscript'
        # Цвет шрифта
        # openpyxl Font.color принимает объект Color.
        color_kwargs = {}
        if 'font_color' in font_db_attrs and font_db_attrs['font_color'] is not None:
            # Если задан прямой RGB цвет (пример: 'FF0000' или 'FFFF0000')
            color_kwargs['rgb'] = font_db_attrs['font_color']
        elif 'font_color_theme' in font_db_attrs and font_db_attrs['font_color_theme'] is not None:
            # Если задана тема цвета
            color_kwargs['theme'] = int(font_db_attrs['font_color_theme'])
            if 'font_color_tint' in font_db_attrs and font_db_attrs['font_color_tint'] is not None:
                color_kwargs['tint'] = float(font_db_attrs['font_color_tint'])
        if color_kwargs:
            try:
                font_kwargs['color'] = Color(**color_kwargs)
            except Exception as e:
                logger.warning(f"[ЭКСПОРТ_СТИЛЕЙ] Ошибка создания Color для Font из атрибутов {color_kwargs}: {e}. Цвет будет стандартным.")
        # Схема шрифта
        if 'font_scheme' in font_db_attrs and font_db_attrs['font_scheme'] is not None:
             font_kwargs['scheme'] = font_db_attrs['font_scheme'] # 'minor', 'major'
        logger.debug(f"[ЭКСПОРТ_СТИЛЕЙ] Подготовлены аргументы Font: {font_kwargs}")
        return Font(**font_kwargs)
    except Exception as e:
        logger.error(f"[ЭКСПОРТ_СТИЛЕЙ] Ошибка создания Font из атрибутов БД {font_db_attrs}: {e}")
        return None
# === КОНЕЦ ИЗМЕНЕНИЙ ===
# === ИЗМЕНЕНО: Функция для создания Fill ===
def _create_openpyxl_fill_from_attrs(fill_db_attrs: Dict[str, Any]) -> Optional[Fill]:
    """Создает объект Fill openpyxl из атрибутов БД."""
    if not fill_db_attrs:
        return None
    try:
        pattern_type = fill_db_attrs.get("fill_pattern_type")
        if pattern_type:
            fill_kwargs = {}
            # Обработка цвета переднего плана
            fg_color_kwargs = {}
            if fill_db_attrs.get("fill_fg_color") is not None:
                fg_color_kwargs['rgb'] = fill_db_attrs["fill_fg_color"]
            elif fill_db_attrs.get("fill_fg_color_theme") is not None:
                fg_color_kwargs['theme'] = int(fill_db_attrs["fill_fg_color_theme"])
                if fill_db_attrs.get("fill_fg_color_tint") is not None:
                    fg_color_kwargs['tint'] = float(fill_db_attrs["fill_fg_color_tint"])
            if fg_color_kwargs:
                try:
                    fill_kwargs['fgColor'] = Color(**fg_color_kwargs)
                except Exception as e:
                    logger.warning(f"[ЭКСПОРТ_СТИЛЕЙ] Ошибка создания fgColor для Fill: {e}")
            # Обработка цвета фона
            bg_color_kwargs = {}
            if fill_db_attrs.get("fill_bg_color") is not None:
                bg_color_kwargs['rgb'] = fill_db_attrs["fill_bg_color"]
            elif fill_db_attrs.get("fill_bg_color_theme") is not None:
                bg_color_kwargs['theme'] = int(fill_db_attrs["fill_bg_color_theme"])
                if fill_db_attrs.get("fill_bg_color_tint") is not None:
                    bg_color_kwargs['tint'] = float(fill_db_attrs["fill_bg_color_tint"])
            if bg_color_kwargs:
                try:
                    fill_kwargs['bgColor'] = Color(**bg_color_kwargs)
                except Exception as e:
                    logger.warning(f"[ЭКСПОРТ_СТИЛЕЙ] Ошибка создания bgColor для Fill: {e}")
            return PatternFill(patternType=pattern_type, **fill_kwargs)
        return None
    except Exception as e:
        logger.error(f"[ЭКСПОРТ_СТИЛЕЙ] Ошибка создания Fill из атрибутов {fill_db_attrs}: {e}")
        return None
# === КОНЕЦ ИЗМЕНЕНИЙ ===
# === ИЗМЕНЕНО: Функция для создания Side ===
def _create_openpyxl_side_from_attrs(side_style: Optional[str], side_color: Optional[str]) -> Optional[Side]:
    """Создает объект Side openpyxl из стиля и цвета."""
    if not side_style and not side_color:
        return None # Отсутствие границы
    try:
        side_kwargs = {}
        if side_style:
            side_kwargs['style'] = side_style
        if side_color:
             # Предполагаем, что цвет передается как RGB строка
             try:
                 side_kwargs['color'] = Color(rgb=side_color)
             except Exception as e:
                 logger.warning(f"[ЭКСПОРТ_СТИЛЕЙ] Ошибка создания Color для Side из '{side_color}': {e}. Используется цвет по умолчанию.")
        return Side(**side_kwargs)
    except Exception as e:
        logger.error(f"[ЭКСПОРТ_СТИЛЕЙ] Ошибка создания Side из style='{side_style}', color='{side_color}': {e}")
        return None # Возвращаем None, что означает отсутствие границы для этой стороны
# === КОНЕЦ ИЗМЕНЕНИЙ ===
# === ИЗМЕНЕНО: Функция для создания Border ===
def _create_openpyxl_border_from_attrs(border_db_attrs: Dict[str, Any]) -> Border:
    """Создает объект Border openpyxl из атрибутов БД."""
    # Даже если атрибутов нет, Border() создаст пустую границу по умолчанию
    try:
        sides = {}
        for side_name in ['left', 'right', 'top', 'bottom', 'diagonal']:
            style_key = f"border_{side_name}_style"
            color_key = f"border_{side_name}_color"
            style = border_db_attrs.get(style_key)
            color = border_db_attrs.get(color_key)
            logger.debug(f"[ЭКСПОРТ_СТИЛЕЙ] Обработка стороны Border '{side_name}': style={style}, color={color}")
            # Создаем Side если хоть что-то задано для этой стороны
            side_obj = _create_openpyxl_side_from_attrs(style, color)
            if side_obj:
                sides[side_name] = side_obj
                logger.debug(f"[ЭКСПОРТ_СТИЛЕЙ] Создана сторона Border '{side_name}': {side_obj}")
            # else: side_obj is None, значит граница для этой стороны отсутствует
        # Добавляем специфичные для диагонали атрибуты
        diagonal_up = border_db_attrs.get("border_diagonal_up")
        diagonal_down = border_db_attrs.get("border_diagonal_down")
        outline = border_db_attrs.get("border_outline")
        border_obj = Border(
            left=sides.get('left'),
            right=sides.get('right'),
            top=sides.get('top'),
            bottom=sides.get('bottom'),
            diagonal=sides.get('diagonal'),
            diagonalUp=diagonal_up,
            diagonalDown=diagonal_down,
            outline=outline
        )
        logger.debug(f"[ЭКСПОРТ_СТИЛЕЙ] Создан объект Border: {border_obj}")
        return border_obj
    except Exception as e:
        logger.error(f"[ЭКСПОРТ_СТИЛЕЙ] Ошибка создания Border из атрибутов БД {border_db_attrs}: {e}")
        # Возвращаем пустую границу в случае ошибки
        return Border()
# === ИЗМЕНЕНО: Функция для создания Alignment ===
def _create_openpyxl_alignment_from_attrs(align_db_attrs: Dict[str, Any]) -> Alignment:
    """Создает объект Alignment openpyxl из атрибутов БД."""
    # Даже если атрибутов нет, Alignment() создаст выравнивание по умолчанию
    try:
        # Создаем словарь аргументов, фильтруя None
        align_kwargs = {k.split('_', 1)[1]: v for k, v in align_db_attrs.items() if k.startswith('alignment_') and v is not None}
        # Особая обработка для BOOLEAN значений, если они хранятся как INTEGER
        # openpyxl обычно принимает bool, но на всякий случай
        bool_keys = ['wrap_text', 'shrink_to_fit', 'justify_last_line']
        for bk in bool_keys:
            if bk in align_kwargs:
                align_kwargs[bk] = bool(align_kwargs[bk])
        alignment_obj = Alignment(**align_kwargs)
        logger.debug(f"[ЭКСПОРТ_СТИЛЕЙ] Создан объект Alignment: {alignment_obj}")
        return alignment_obj
    except Exception as e:
        logger.error(f"[ЭКСПОРТ_СТИЛЕЙ] Ошибка создания Alignment из атрибутов {align_db_attrs}: {e}")
        return Alignment()
# === КОНЕЦ ИЗМЕНЕНИЙ ===
# === ИЗМЕНЕНО: Функция для создания Protection ===
def _create_openpyxl_protection_from_attrs(prot_db_attrs: Dict[str, Any]) -> Protection:
    """Создает объект Protection openpyxl из атрибутов БД."""
    # Даже если атрибутов нет, Protection() создаст защиту по умолчанию (обычно unlocked=False, hidden=False)
    try:
        # Создаем словарь аргументов, фильтруя None
        # Приводим BOOLEAN значения к INTEGER/bool если нужно, хотя openpyxl обычно принимает bool
        prot_kwargs = {}
        for k, v in prot_db_attrs.items():
            if k.startswith('protection_') and v is not None:
                attr_name = k.split('_', 1)[1]
                # openpyxl Protection ожидает bool для locked и hidden
                if attr_name in ['locked', 'hidden']:
                    prot_kwargs[attr_name] = bool(v)
                else:
                    prot_kwargs[attr_name] = v
        protection_obj = Protection(**prot_kwargs)
        logger.debug(f"[ЭКСПОРТ_СТИЛЕЙ] Создан объект Protection: {protection_obj}")
        return protection_obj
    except Exception as e:
        logger.error(f"[ЭКСПОРТ_СТИЛЕЙ] Ошибка создания Protection из атрибутов {prot_db_attrs}: {e}")
        return Protection()
# === КОНЕЦ ИЗМЕНЕНИЙ ===
# === ИЗМЕНЕНО: Функция для создания NamedStyle ===
def _create_named_style_from_style_attrs(style_db_attrs: Dict[str, Any], style_name: str) -> Optional[NamedStyle]:
    """
    Создает именованный стиль openpyxl из атрибутов стиля БД.
    Args:
        style_db_attrs (Dict[str, Any]): Все атрибуты стиля, загруженные из БД.
                                        Ожидается, что ключи включают префиксы (font_*, fill_*, ...).
        style_name (str): Имя создаваемого именованного стиля.
    Returns:
        Optional[NamedStyle]: Созданный объект NamedStyle или None в случае ошибки.
    """
    logger.debug(f"[ЭКСПОРТ_СТИЛЕЙ] Создание именованного стиля '{style_name}' из атрибутов БД.")
    try:
        named_style = NamedStyle(name=style_name)
        # Извлекаем атрибуты для каждого компонента стиля
        font_attrs = {k.split('_', 1)[1]: v for k, v in style_db_attrs.items() if k.startswith('font_')}
        fill_attrs = {k.split('_', 1)[1]: v for k, v in style_db_attrs.items() if k.startswith('fill_')}
        border_attrs = {k.split('_', 1)[1]: v for k, v in style_db_attrs.items() if k.startswith('border_')}
        align_attrs = {k.split('_', 1)[1]: v for k, v in style_db_attrs.items() if k.startswith('alignment_')}
        prot_attrs = {k.split('_', 1)[1]: v for k, v in style_db_attrs.items() if k.startswith('protection_')}
        logger.debug(f"[ЭКСПОРТ_СТИЛЕЙ] Атрибуты Font для стиля '{style_name}': {font_attrs}")
        logger.debug(f"[ЭКСПОРТ_СТИЛЕЙ] Атрибуты Fill для стиля '{style_name}': {fill_attrs}")
        logger.debug(f"[ЭКСПОРТ_СТИЛЕЙ] Атрибуты Border для стиля '{style_name}': {border_attrs}")
        logger.debug(f"[ЭКСПОРТ_СТИЛЕЙ] Атрибуты Alignment для стиля '{style_name}': {align_attrs}")
        logger.debug(f"[ЭКСПОРТ_СТИЛЕЙ] Атрибуты Protection для стиля '{style_name}': {prot_attrs}")
        # Создаем и применяем компоненты
        if font_attrs:
            font_obj = _create_openpyxl_font_from_attrs(font_attrs)
            if font_obj:
                named_style.font = font_obj
                logger.debug(f"[ЭКСПОРТ_СТИЛЕЙ] Font применен к стилю '{style_name}'")
        if fill_attrs:
            fill_obj = _create_openpyxl_fill_from_attrs(fill_attrs)
            if fill_obj:
                named_style.fill = fill_obj
                logger.debug(f"[ЭКСПОРТ_СТИЛЕЙ] Fill применен к стилю '{style_name}'")
        # Border всегда создается, даже пустой
        border_obj = _create_openpyxl_border_from_attrs(border_attrs)
        named_style.border = border_obj
        logger.debug(f"[ЭКСПОРТ_СТИЛЕЙ] Border применен к стилю '{style_name}'")
        # Alignment всегда создается, даже пустой
        align_obj = _create_openpyxl_alignment_from_attrs(align_attrs)
        named_style.alignment = align_obj
        logger.debug(f"[ЭКСПОРТ_СТИЛЕЙ] Alignment применен к стилю '{style_name}'")
        # Protection всегда создается, даже пустой
        prot_obj = _create_openpyxl_protection_from_attrs(prot_attrs)
        named_style.protection = prot_obj
        logger.debug(f"[ЭКСПОРТ_СТИЛЕЙ] Protection применен к стилю '{style_name}'")
        # Атрибуты самого стиля (те, что не входят в подкомпоненты)
        # num_fmt_id, xf_id, quote_prefix обычно обрабатываются на уровне ячейки или книги
        # named_style.number_format = ... # TODO: если нужно
        logger.debug(f"[ЭКСПОРТ_СТИЛЕЙ] Именованный стиль '{style_name}' успешно создан.")
        return named_style
    except Exception as e:
        logger.error(f"[ЭКСПОРТ_СТИЛЕЙ] Ошибка создания именованного стиля '{style_name}': {e}", exc_info=True)
        return None
# === КОНЕЦ ИЗМЕНЕНИЙ ===
# === НОВОЕ: Функция для получения имен существующих стилей ===
def _get_existing_style_names(workbook: Workbook) -> Set[str]:
    """
    Получает множество имен существующих именованных стилей в книге.
    Args:
        workbook (Workbook): Объект рабочей книги openpyxl.
    Returns:
        Set[str]: Множество имен существующих стилей.
    """
    try:
        # === ИСПРАВЛЕНО: Корректная обработка workbook.named_styles ===
        # workbook.named_styles возвращает список имен (строк) существующих стилей
        # НЕ нужно итерироваться по нему как по списку объектов с атрибутом 'name'
        existing_style_names = set(workbook.named_styles)
        logger.debug(f"[ЭКСПОРТ_СТИЛЕЙ] Получены имена существующих стилей из workbook.named_styles: {existing_style_names}")
        return existing_style_names
        # === КОНЕЦ ИСПРАВЛЕНИЙ ===
    except Exception as e:
        logger.error(f"[ЭКСПОРТ_СТИЛЕЙ] Ошибка при получении имен существующих стилей: {e}")
        return set() # Возвращаем пустое множество в случае ошибки
# === КОНЕЦ НОВОГО ===
# === ИЗМЕНЕНО: Основная функция экспорта стилей ===
def export_sheet_styles(ws: Worksheet, styled_ranges_data: List[Dict[str, Any]], workbook: Workbook) -> None:
    """
    Экспортирует стили ячеек листа.
    Args:
        ws (Worksheet): Лист Excel, на который применяются стили.
        styled_ranges_data (List[Dict[str, Any]]): Список словарей с информацией о стилях.
                                                   Ожидается формат:
                                                   [{'style_attributes': {...}, 'range_address': 'A1'}, ...]
                                                   или
                                                   [{'font_b': 1, ..., 'range_address': 'A1:B2'}, ...]
        workbook (Workbook): Рабочая книга для добавления именованных стилей.
    """
    logger.info(f"[ЭКСПОРТ_СТИЛЕЙ] === Начало экспорта стилей для листа '{ws.title}' ===")
    logger.debug(f"[ЭКСПОРТ_СТИЛЕЙ] Получено {len(styled_ranges_data)} записей о стилях.")
    if not styled_ranges_data:
        logger.info("[ЭКСПОРТ_СТИЛЕЙ] Нет стилей для экспорта.")
        return
    applied_styles_count = 0
    # === ИСПРАВЛЕНО: Безопасное получение имен существующих стилей ===
    # Используем новую функцию для получения множества имен
    existing_style_names: Set[str] = _get_existing_style_names(workbook)
    logger.debug(f"[ЭКСПОРТ_СТИЛЕЙ] Известные имена стилей в книге: {existing_style_names}")
    # === КОНЕЦ ИСПРАВЛЕНИЙ ===
    for i, style_range_info in enumerate(styled_ranges_data):
        logger.debug(f"[ЭКСПОРТ_СТИЛЕЙ] Обработка стиля {i+1}: {style_range_info}")
        try:
            # Предполагаем, что style_range_info - это словарь
            # с ключами, включая атрибуты стиля и 'range_address'
            # Извлекаем адрес диапазона
            range_addr = style_range_info.get("range_address", "")
            if not range_addr:
                 logger.warning(f"[ЭКСПОРТ_СТИЛЕЙ] Пропущена запись стиля {i+1} из-за отсутствия 'range_address': {style_range_info}")
                 continue
            # Создаем уникальное имя для стиля на основе его атрибутов
            # Исключаем range_address из хеширования
            attrs_for_hash = {k: v for k, v in style_range_info.items() if k != "range_address"}
            # Простой способ генерации уникального имени, можно улучшить
            style_name = f"Style_{abs(hash(str(sorted(attrs_for_hash.items()))) % 10000000)}"
            logger.debug(f"[ЭКСПОРТ_СТИЛЕЙ] Сгенерировано имя стиля: {style_name}")
            # === ИСПРАВЛЕНО: Проверяем существование стиля и добавляем его при необходимости ===
            # Проверяем по множеству existing_style_names
            if style_name not in existing_style_names:
                # Создаем именованный стиль
                named_style = _create_named_style_from_style_attrs(style_range_info, style_name)
                if named_style:
                    # Добавляем стиль в книгу
                    try:
                        workbook.add_named_style(named_style)
                        logger.debug(f"[ЭКСПОРТ_СТИЛЕЙ] Добавлен новый именованный стиль '{style_name}' в книгу.")
                        # Обновляем множество известных имен
                        existing_style_names.add(style_name)
                    except Exception as add_style_e:
                        logger.error(f"[ЭКСПОРТ_СТИЛЕЙ] Ошибка добавления стиля '{style_name}' в книгу: {add_style_e}")
                        # Продолжаем, возможно стиль уже добавлен или будет добавлен позже
                        # Проверим еще раз
                        if style_name not in _get_existing_style_names(workbook): # <-- Используем исправленную функцию
                             logger.error(f"[ЭКСПОРТ_СТИЛЕЙ] Стиль '{style_name}' не добавлен и не найден. Пропущен.")
                             continue
                        else:
                            logger.info(f"[ЭКСПОРТ_СТИЛЕЙ] Стиль '{style_name}' найден в книге после ошибки добавления.")
                else:
                    logger.error(f"[ЭКСПОРТ_СТИЛЕЙ] Не удалось создать именованный стиль для {attrs_for_hash}")
                    continue
            else:
                 logger.debug(f"[ЭКСПОРТ_СТИЛЕЙ] Именованный стиль '{style_name}' уже существует в книге.")
            # === КОНЕЦ ИЗМЕНЕНИЙ ===
            # Применяем стиль к диапазону на листе
            try:
                cell_range = ws[range_addr]
                cells_to_style: List[Cell] = []
                # === ИЗМЕНЕНО: Упрощенная и более надежная итерация ===
                if isinstance(cell_range, Cell):
                    # Это одиночная ячейка
                    cells_to_style = [cell_range]
                elif hasattr(cell_range, '__iter__'):
                    # Это итерируемый объект (например, tuple из CellRange или Cell)
                    for item in cell_range:
                        if isinstance(item, Cell):
                            cells_to_style.append(item)
                        elif hasattr(item, '__iter__'):
                            # Это CellRange, итерируемся по ячейкам внутри
                            for cell in item:
                                cells_to_style.append(cell)
                        else:
                            logger.warning(f"[ЭКСПОРТ_СТИЛЕЙ] Неожиданный тип элемента в CellRange: {type(item)}")
                else:
                    # Теоретически не должно произойти, так как ws[addr] всегда возвращает Cell или CellRange
                    logger.warning(f"[ЭКСПОРТ_СТИЛЕЙ] Неожиданный тип для диапазона '{range_addr}': {type(cell_range)}. Пропущен.")
                    continue
                # === КОНЕЦ ИЗМЕНЕНИЙ ===
                logger.debug(f"[ЭКСПОРТ_СТИЛЕЙ] Применение стиля '{style_name}' к {len(cells_to_style)} ячейкам в диапазоне '{range_addr}'.")
                for cell in cells_to_style:
                    try:
                        # === ИСПРАВЛЕНО: Применяем стиль по имени ===
                        # Проверяем, что стиль существует в книге перед применением
                        # Используем внутренний словарь для более эффективной проверки
                        if style_name in workbook._named_styles:
                            try:
                                # Применяем стиль к ячейке, используя имя именованного стиля
                                # Это публичный способ связать ячейку с именованным стилем
                                cell.style = style_name
                                logger.debug(f"[ЭКСПОРТ_СТИЛЕЙ] Стиль '{style_name}' применен к ячейке {cell.coordinate}.")
                            except Exception as apply_cell_style_e:
                                logger.error(f"[ЭКСПОРТ_СТИЛЕЙ] Ошибка применения стиля '{style_name}' к ячейке {cell.coordinate}: {apply_cell_style_e}")
                        else:
                            logger.warning(f"[ЭКСПОРТ_СТИЛЕЙ] Стиль '{style_name}' не найден в книге. Не удалось применить к ячейке {cell.coordinate}.")
                        # === КОНЕЦ ИСПРАВЛЕНИЙ ===
                    except Exception as apply_cell_style_e:
                        logger.error(f"[ЭКСПОРТ_СТИЛЕЙ] Ошибка применения стиля '{style_name}' к ячейке {cell.coordinate}: {apply_cell_style_e}")
                applied_styles_count += 1
                logger.debug(f"[ЭКСПОРТ_СТИЛЕЙ] Стиль '{style_name}' успешно применен к диапазону '{range_addr}'.")
            except Exception as apply_e:
                logger.error(f"[ЭКСПОРТ_СТИЛЕЙ] Ошибка применения стиля '{style_name}' к диапазону '{range_addr}': {apply_e}")
        except Exception as e:
            logger.error(f"[ЭКСПОРТ_СТИЛЕЙ] Ошибка обработки стиля {i+1}: {e}", exc_info=True)
    logger.info(f"[ЭКСПОРТ_СТИЛЕЙ] === Конец экспорта стилей для листа '{ws.title}'. Применено {applied_styles_count} стилей. ===")
# === КОНЕЦ ИЗМЕНЕНИЙ ===
# - ТОЧКА ВХОДА ДЛЯ ТЕСТИРОВАНИЯ -
if __name__ == "__main__":
    # Простой тест подключения и инициализации
    print("--- ТЕСТ ХРАНИЛИЩА ---")
    # Определяем путь к тестовой БД относительно корня проекта
    test_db_path = project_root / "data" / "test_db.sqlite"
    print(f"Путь к тестовой БД: {test_db_path}")
    try:
        # === ИСПРАВЛЕНО: Импорт ProjectDBStorage ===
        from src.storage.base import ProjectDBStorage # <-- Добавлен импорт
        # === КОНЕЦ ИСПРАВЛЕНИЙ ===
        storage = ProjectDBStorage(str(test_db_path))
        storage.connect()
        print("Подключение к тестовой БД установлено и схема инициализирована.")
        storage.disconnect()
        print("Подключение закрыто.")
        # Пытаемся удалить тестовый файл БД
        if test_db_path.exists():
            test_db_path.unlink()
            print("Тестовый файл БД удален.")
    except Exception as e:
        print(f"Ошибка при тестировании хранилища: {e}")
