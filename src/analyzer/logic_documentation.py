# src/analyzer/logic_documentation.py

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
# --- ИСПРАВЛЕНИЕ: Импортируем Cell напрямую с псевдонимом ---
# Это должно помочь Pylance понять тип, совместимый с _CellOrMergedCell
from openpyxl.cell.cell import Cell as OpenPyxlCell
# Для аннотаций типов
from typing import Dict, Any, List, Optional
import logging
import json
import re

# Импортируем logger из utils
from src.utils.logger import get_logger

logger = get_logger(__name__)

# --- Вспомогательные функции для сериализации сложных объектов ---

# --- Вспомогательная функция для извлечения RGB ---
def _get_rgb_string(color_obj):
    """
    Извлекает строковое RGB-представление из объекта цвета openpyxl.
    Поддерживает RGB, Color, ThemeColor.
    Возвращает строку RGB (например, 'FF0000') или None, если цвет не может быть представлен в RGB.
    """
    if color_obj is None:
        return None

    # Проверяем, является ли объект экземпляром RGB (часто используется для цветов)
    # или имеет атрибут rgb, как у Color
    if hasattr(color_obj, 'rgb'):
        # Это может быть Color или RGB
        # RGB-объекты также имеют атрибут rgb, но он может быть не строкой
        # Попробуем получить .rgb
        rgb_val = getattr(color_obj, 'rgb', None)
        if isinstance(rgb_val, str):
            # Если rgb - строка, это и есть нужное нам значение
            return rgb_val
        # Если rgb не строка, возможно, это ThemeColor или что-то другое
        # Попробуем строковое представление
        return str(color_obj)

    # Если у объекта нет атрибута rgb, но он сам является строкой (например, RGB),
    # используем строковое представление
    if isinstance(color_obj, str):
        return color_obj

    # Попробуем строковое представление для других типов (например, ThemeColor)
    # ThemeColor может иметь атрибуты типа tint, theme, но не rgb
    # Если строковое представление не даст полезного результата (например, '<ThemeColor...>'),
    # можно вернуть None или обработать особым образом.
    # Для ThemeColor, если он не преобразуется в RGB, возвращаем None.
    if hasattr(color_obj, 'type') and getattr(color_obj, 'type') == 'theme':
         # ThemeColor не может быть представлен как простой RGB
         return None

    # Попробуем строковое представление как последнюю надежду
    str_repr = str(color_obj)
    # Проверим, похоже ли строковое представление на RGB (6 или 8 шестнадцатеричных символов, возможно, с '0x' или '#')
    if re.fullmatch(r'[A-Fa-f0-9]{6}|[A-Fa-f0-9]{8}', str_repr):
        return str_repr

    # Если ничего не подошло, возвращаем None
    return None


def _serialize_style(cell: OpenPyxlCell | Any) -> Dict[str, Any]:
    """
    Сериализует атрибуты стиля ячейки openpyxl в словарь.
    Этот словарь будет сериализован в JSON в storage/styles.py.
    Структура должна соответствовать ожиданиям _convert_style_to_xlsxwriter_format
    в excel_exporter.py.
    """
    # --- ИСПРАВЛЕНИЕ: Pylance теперь должен понимать, что у cell (OpenPyxlCell | Any) есть атрибуты font, fill и т.д.
    # Проверки hasattr остаются для безопасности во время выполнения.
    style_dict = {}

    # --- Шрифт ---
    if hasattr(cell, 'font') and cell.font:
        font_dict = {}
        if cell.font.name: font_dict['name'] = cell.font.name
        if cell.font.sz: font_dict['sz'] = cell.font.sz # Размер
        if cell.font.b: font_dict['b'] = cell.font.b # Жирный
        if cell.font.i: font_dict['i'] = cell.font.i # Курсив
        if cell.font.color and _get_rgb_string(cell.font.color):
            font_dict['color'] = {'rgb': _get_rgb_string(cell.font.color)}
        # ... другие атрибуты шрифта (underline, vertAlign и т.д.)
        if font_dict:
            style_dict['font'] = font_dict

    # --- Заливка ---
    if hasattr(cell, 'fill') and cell.fill:
        fill_dict = {}
        # openpyxl.fill.PatternFill или openpyxl.fill.GradientFill
        if hasattr(cell.fill, 'patternType'):
             fill_dict['patternType'] = cell.fill.patternType
        if hasattr(cell.fill, 'fgColor') and cell.fill.fgColor and _get_rgb_string(cell.fill.fgColor):
             fill_dict['fgColor'] = {'rgb': _get_rgb_string(cell.fill.fgColor)}
        if hasattr(cell.fill, 'bgColor') and cell.fill.bgColor and _get_rgb_string(cell.fill.bgColor):
             fill_dict['bgColor'] = {'rgb': _get_rgb_string(cell.fill.bgColor)}
        # ... обработка GradientFill
        if fill_dict:
             style_dict['fill'] = fill_dict

    # --- Границы ---
    if hasattr(cell, 'border') and cell.border:
        border_dict = {}
        # openpyxl.styles.borders.Border
        for side_name in ['left', 'right', 'top', 'bottom']: # 'diagonal' ?
            side_obj = getattr(cell.border, side_name, None)
            if side_obj and (side_obj.style or (side_obj.color and _get_rgb_string(side_obj.color))):
                side_dict = {}
                if side_obj.style: side_dict['style'] = side_obj.style
                if side_obj.color and _get_rgb_string(side_obj.color):
                    side_dict['color'] = {'rgb': _get_rgb_string(side_obj.color)}
                border_dict[side_name] = side_dict
        if border_dict:
            style_dict['border'] = border_dict

    # --- Выравнивание ---
    if hasattr(cell, 'alignment') and cell.alignment:
        alignment_dict = {}
        # openpyxl.styles.alignment.Alignment
        if cell.alignment.horizontal: alignment_dict['horizontal'] = cell.alignment.horizontal
        if cell.alignment.vertical: alignment_dict['vertical'] = cell.alignment.vertical
        if cell.alignment.wrapText is not None: alignment_dict['wrapText'] = cell.alignment.wrapText
        if cell.alignment.textRotation is not None: alignment_dict['textRotation'] = cell.alignment.textRotation
        # ... другие атрибуты (shrinkToFit, indent и т.д.)
        if alignment_dict:
            style_dict['alignment'] = alignment_dict

    # --- Числовой формат ---
    if hasattr(cell, 'number_format') and cell.number_format:
        style_dict['number_format'] = cell.number_format

    # --- Защита ---
    if hasattr(cell, 'protection') and cell.protection:
        protection_dict = {}
        # openpyxl.styles.protection.Protection
        if cell.protection.locked is not None: protection_dict['locked'] = cell.protection.locked
        if cell.protection.hidden is not None: protection_dict['hidden'] = cell.protection.hidden
        if protection_dict:
            style_dict['protection'] = protection_dict

    # logger.debug(f"Сериализован стиль для ячейки {cell.coordinate}: {list(style_dict.keys())}")
    return style_dict


def _serialize_chart(chart_obj) -> Dict[str, Any]:
    """
    Сериализует объект диаграммы openpyxl.
    Включает тип, данные, заголовок, позицию и размер.
    """
    try:
        chart_data = {}
        chart_data['type'] = type(chart_obj).__name__ # 'BarChart', 'LineChart' и т.д.

        # --- Извлечение позиции и размера ---
        # Привязка диаграммы к листу (anchor) определяет её позицию и размер
        # Инициализируем переменные, которые могут не быть установлены, если не будет 'anchor'
        position_info = None
        width_emu = None
        height_emu = None

        if hasattr(chart_obj, 'anchor'): 
            anchor = chart_obj.anchor
            logger.debug(f"[ДИАГРАММА] Найден anchor типа: {type(anchor).__name__} для {chart_data['type']}")
            
            # Проверяем тип привязки
            if hasattr(anchor, '_from'):
                # Это может быть OneCellAnchor или TwoCellAnchor
                from_cell = anchor._from
                to_cell = getattr(anchor, 'to', None) # 'to' может не быть у OneCellAnchor
                
                position_info = {
                    "from_col": from_cell.col,
                    "from_row": from_cell.row,
                    "from_col_offset": from_cell.colOff,
                    "from_row_offset": from_cell.rowOff,
                }
                
                # Инициализируем переменные для размеров
                width_emu = None
                height_emu = None
                
                if to_cell:
                    # TwoCellAnchor: позиция определяется двумя ячейками
                    position_info.update({
                        "to_col": to_cell.col,
                        "to_row": to_cell.row,
                        "to_col_offset": to_cell.colOff,
                        "to_row_offset": to_cell.rowOff,
                    })
                    # Для TwoCellAnchor размер можно вычислить из разности координат
                    # Ширина = (to_col - from_col) * ширина_ячейки + (to_col_offset - from_col_offset)
                    # Высота = (to_row - from_row) * высота_ячейки + (to_row_offset - from_row_offset)
                    # Это сложно, так как ширина/высота ячеек переменна. 
                    # xlsxwriter использует from/to напрямую, поэтому сохраняем их.
                    # Но для полноты информации попробуем вычислить размеры в EMU.
                    # Приблизительное вычисление (не учитывает ширину/высоту отдельных ячеек):
                    # Ширина в EMU = (to_col - from_col) * 640000 + (to_col_offset - from_col_offset)
                    # Высота в EMU = (to_row - from_row) * 640000 + (to_row_offset - from_row_offset)
                    # 640000 EMU = 1/914400 дюйма * 96 DPI * 6 (приблизительная ширина/высота ячейки в символах)
                    # Это грубое приближение. Лучше использовать anchor.ext, если он есть.
                    try:
                        ext = getattr(anchor, 'ext', None)
                        if ext:
                            width_emu = ext.width
                            height_emu = ext.height
                        else:
                            # Грубое приближение, если ext недоступен
                            width_emu = (to_cell.col - from_cell.col) * 640000 + (to_cell.colOff - from_cell.colOff)
                            height_emu = (to_cell.row - from_cell.row) * 640000 + (to_cell.rowOff - from_cell.rowOff)
                    except Exception as calc_e:
                        logger.warning(f"[ДИАГРАММА] Не удалось вычислить размеры для TwoCellAnchor: {calc_e}")
                        # Если вычисление не удалось, оставляем width_emu/height_emu как None
                else:
                    # OneCellAnchor: позиция определяется одной ячейкой и размером
                    # Размеры берём из chart_obj.width/height или anchor.ext
                    ext = getattr(anchor, 'ext', None)
                    width_emu = getattr(chart_obj, 'width', None) or (ext.width if ext else None)
                    height_emu = getattr(chart_obj, 'height', None) or (ext.height if ext else None)
                
                # Сохраняем позицию
                chart_data['position'] = position_info
                
                # Сохраняем размеры в EMU, если они были извлечены
                if width_emu is not None:
                    chart_data['width_emu'] = width_emu
                if height_emu is not None:
                    chart_data['height_emu'] = height_emu
            else:
                logger.warning(f"[ДИАГРАММА] Anchor не имеет атрибута _from: {type(anchor).__name__}")
        else:
            logger.warning(f"[ДИАГРАММА] У диаграммы {chart_data['type']} нет атрибута 'anchor'")
        # --- Конец извлечения позиции и размера ---
        # Сохраняем позицию и размеры в chart_data, если они были извлечены
        # (переменные уже инициализированы, проблем с 'possibly unbound' не будет)
        if position_info:
            chart_data['position'] = position_info
        if width_emu is not None:
            chart_data['width_emu'] = width_emu
        if height_emu is not None:
            chart_data['height_emu'] = height_emu

        # Пример: сохранение ссылок на данные
        if hasattr(chart_obj, 'ser') and chart_obj.ser:
            series_data = [] # Инициализация списка серий
            for idx, s in enumerate(chart_obj.ser):
                ser_dict = {} # Инициализация словаря для каждой серии
                # Сохраняем адреса диапазонов данных
                if hasattr(s, 'val') and s.val and hasattr(s.val, 'numRef') and s.val.numRef:
                    ser_dict['val_range'] = s.val.numRef.f # Строка формулы диапазона значений
                if hasattr(s, 'cat') and s.cat and hasattr(s.cat, 'strRef') and s.cat.strRef:
                    ser_dict['cat_range'] = s.cat.strRef.f # Строка формулы диапазона категорий
                # ... другие атрибуты серии (название, цвета и т.д.)
                series_data.append(ser_dict) # Добавляем словарь серии в список
            # После цикла устанавливаем 'series' в chart_data
            chart_data['series'] = series_data
        else:
            # Если серий нет, устанавливаем пустой список
            chart_data['series'] = []

        # --- Улучшенное извлечение заголовка ---
        if hasattr(chart_obj, 'title') and chart_obj.title:
            title_obj = chart_obj.title
            title_text = None
            
            # Проверяем, есть ли текстовый заголовок (tx)
            if hasattr(title_obj, 'tx') and title_obj.tx:
                tx_obj = title_obj.tx
                
                # Проверяем Rich Text
                if hasattr(tx_obj, 'rich') and tx_obj.rich:
                    try:
                        # Берем первый параграф и первый run
                        if (hasattr(tx_obj.rich, 'p') and tx_obj.rich.p and 
                            len(tx_obj.rich.p) > 0 and tx_obj.rich.p[0] and
                            hasattr(tx_obj.rich.p[0], 'r') and tx_obj.rich.p[0].r and
                            len(tx_obj.rich.p[0].r) > 0):
                            
                            first_run = tx_obj.rich.p[0].r[0]
                            # Получаем текст из run (может быть пустым "")
                            title_text = getattr(first_run, 't', "") 
                            
                    except (AttributeError, IndexError) as ae:
                        logger.debug(f"[ДИАГРАММА] Не удалось извлечь заголовок из rich text: {ae}")
                
                # Проверяем ссылку на ячейку
                elif hasattr(tx_obj, 'strRef') and tx_obj.strRef:
                    chart_data['title_ref'] = tx_obj.strRef.f
            
            # Если текст был извлечен (даже если он ""), сохраняем его
            if title_text is not None:
                chart_data['title'] = title_text
            # --- Конец улучшенного извлечения заголовка ---

            # --- Улучшенное извлечение легенды ---
            if hasattr(chart_obj, 'legend') and chart_obj.legend:
                legend_obj = chart_obj.legend
                legend_data = {}
                
                # Извлекаем позицию легенды, если она задана
                if hasattr(legend_obj, 'position') and legend_obj.position:
                    legend_data['position'] = legend_obj.position
                
                # Можно добавить другие атрибуты легенды, если потребуется
                # Например, видимость (хотя если объект legend существует, она, скорее всего, True)
                # if hasattr(legend_obj, 'show_legend') and legend_obj.show_legend is not None:
                #     legend_data['show_legend'] = legend_obj.show_legend
                
                # Добавляем извлечённые данные легенды в основной словарь chart_data
                chart_data['legend'] = legend_data
                logger.debug(f"[ДИАГРАММА] Извлечена информация о легенде: {legend_data}")
            else:
                logger.debug(f"[ДИАГРАММА] У диаграммы нет объекта легенды или он пуст.")
            # --- Конец улучшенного извлечения легенды ---

        logger.debug(f"Сериализована диаграмма типа {chart_data.get('type', 'Unknown')}")
        return chart_data

    except Exception as e:
        logger.error(f"Ошибка при сериализации диаграммы: {e}", exc_info=True)
        return {}

# --- Основная функция анализа ---

def analyze_excel_file(file_path: str) -> Dict[str, Any]:
    """
    Основная функция для анализа Excel-файла.
    Извлекает структуру, данные, формулы, стили, диаграммы и другую информацию.
    Возвращает словарь с результатами анализа, готовый для передачи в storage.

    Args:
        file_path (str): Путь к анализируемому .xlsx файлу.

    Returns:
        Dict[str, Any]: Словарь с результатами анализа.
    """
    logger.info(f"Начало анализа Excel-файла: {file_path}")

    try:
        # Открываем книгу Excel
        workbook = openpyxl.load_workbook(file_path, data_only=False) # data_only=False для получения формул
        logger.debug(f"Книга '{file_path}' успешно открыта.")

        analysis_results = {
            "project_name": file_path.split("/")[-1].split(".")[0], # Имя проекта из имени файла
            "file_path": file_path,
            "sheets": []
        }

        # Итерируемся по всем листам в книге
        for sheet_name in workbook.sheetnames:
            logger.info(f"Анализ листа: {sheet_name}")
            sheet: Worksheet = workbook[sheet_name]

            sheet_data = {
                "name": sheet_name,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "raw_data": [],
                "formulas": [],
                "styles": [], # Будет содержать {'range_address': str, 'style_attributes': str (JSON)}
                "charts": [], # Будет содержать {'chart_data': dict или str}
                "merged_cells": [] # Список строк адресов объединенных ячеек
            }

            # --- 1. Извлечение "сырых данных" ---
            logger.debug(f"Извлечение сырых данных с листа '{sheet_name}'...")
            for row in sheet.iter_rows(values_only=False): # values_only=False, чтобы получить объекты Cell
                for cell in row:
                    if cell.value is not None or cell.data_type == 'f': # Сохраняем и данные, и формулы
                        data_item = {
                            "cell_address": cell.coordinate,
                            "value": cell.value,
                            # "value_type": type(cell.value).__name__ # Может быть полезно
                        }
                        sheet_data["raw_data"].append(data_item)

            # --- 2. Извлечение формул ---
            logger.debug(f"Извлечение формул с листа '{sheet_name}'...")
            for row in sheet.iter_rows(values_only=False):
                for cell in row:
                    # Проверим, является ли значение строкой, начинающейся с '='
                    # Или используем data_type для надежности
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                         sheet_data["formulas"].append({
                             "cell_address": cell.coordinate,
                             "formula": cell.value # Сохраняем формулу как есть, включая '='
                         })
                    # ВАЖНО: openpyxl также имеет cell.formula, но это может быть не то же самое.
                    # Уточнение: cell.value при data_only=False содержит формулу.

            # --- 3. Извлечение стилей ---
            logger.debug(f"Извлечение и группировка стилей с листа '{sheet_name}'...")
            # Простой подход: для каждой ячейки сохраняем её стиль.
            # Более сложный (эффективный): группировать ячейки с одинаковыми стилями в диапазоны.
            # Пока используем простой подход для MVP.

            # Словарь для хранения уникальных стилей и их диапазонов
            style_ranges_map: Dict[str, List[str]] = {} # ключ - сериализованный стиль, значение - список адресов

            for row in sheet.iter_rows(values_only=False):
                for cell in row:
                    # Проверяем, есть ли у ячейки не-дефолтный стиль
                    # Это можно сделать, сравнивая с дефолтным стилем, но проще сериализовать всегда
                    # и потом в storage/styles.py решать, нужно ли его сохранять.

                    # --- ИСПРАВЛЕНИЕ: Передаем cell в _serialize_style, аннотированную как OpenPyxlCell | Any ---
                    # iter_rows возвращает _CellOrMergedCell. Pylance должен принять OpenPyxlCell | Any.
                    style_dict = _serialize_style(cell)
                    # --- КОНЕЦ ИСПРАВЛЕНИЯ ---
                    if style_dict: # Если стиль не пустой
                        style_json = json.dumps(style_dict, sort_keys=True) # Сериализуем в JSON и сортируем ключи для надежного хеширования
                        coord = cell.coordinate

                        if style_json in style_ranges_map:
                            style_ranges_map[style_json].append(coord)
                        else:
                            style_ranges_map[style_json] = [coord]

            # Преобразуем карту стилей в формат, ожидаемый storage
            for style_json, cell_addresses in style_ranges_map.items():
                # Для упрощения, будем создавать отдельную запись для каждой ячейки
                # В будущем можно реализовать группировку в диапазоны (A1:A10, B1:D1 и т.д.)
                # Это требует сложной логики группировки.
                for address in cell_addresses:
                     sheet_data["styles"].append({
                         "range_address": address, # Пока каждая ячейка отдельно
                         "style_attributes": style_json # Строка JSON
                     })
                # TODO: Реализовать группировку адресов в диапазоны
                # sheet_data["styles"].append({
                #     "range_address": _group_addresses(cell_addresses), # Функция группировки
                #     "style_attributes": style_json # Строка JSON
                # })

            # --- 4. Извлечение диаграмм ---
            logger.debug(f"Извлечение диаграмм с листа '{sheet_name}'...")
            # Диаграммы находятся в sheet._charts
            # --- ИСПРАВЛЕНИЕ: Оборачиваем доступ к _charts в try/except и добавляем # type: ignore ---
            try:
                # Атрибут _charts является внутренним. Доступ к нему может быть ненадежным.
                # Используем # type: ignore[attr-defined] для подавления предупреждения Pylance.
                charts_list = sheet._charts # type: ignore[attr-defined]
                for chart_obj in charts_list:
                     chart_data = _serialize_chart(chart_obj)
                     if chart_data:
                         # storage ожидает 'chart_data' как сериализованный объект
                         sheet_data["charts"].append({
                             "chart_data": chart_data # Это будет словарь, storage должен его сериализовать при сохранении
                             # Если нужно сохранить как JSON сразу:
                             # "chart_data": json.dumps(chart_data, ensure_ascii=False)
                         })
            except AttributeError as ae:
                logger.warning(f"Не удалось получить доступ к диаграммам листа '{sheet_name}' через _charts: {ae}")
            except Exception as e:
                logger.error(f"Ошибка при извлечении диаграмм с листа '{sheet_name}': {e}", exc_info=True)
            # --- КОНЕЦ ИСПРАВЛЕНИЯ ---


            # --- 5. Извлечение объединенных ячеек ---
            logger.debug(f"Извлечение объединенных ячеек с листа '{sheet_name}'...")
            logger.debug(f"Объект merged_cells: {sheet.merged_cells}")
            logger.debug(f"Атрибут ranges объекта merged_cells: {getattr(sheet.merged_cells, 'ranges', 'NO_RANGES_ATTR')}")
            merged_ranges = getattr(sheet.merged_cells, 'ranges', [])
            logger.debug(f"Итерация по ranges...")
            merged_count = 0
            for merged_cell_range in merged_ranges:
                merged_count += 1
                logger.debug(f"Найден объединенный диапазон: {merged_cell_range} (Тип: {type(merged_cell_range)})")
                # merged_cell_range это openpyxl.utils.cell_range.CellRange
                range_str = str(merged_cell_range)
                logger.debug(f"Преобразован в строку: {range_str}")
                sheet_data["merged_cells"].append(range_str) # Преобразуем в строку адреса диапазона
                logger.debug(f"Добавлен в sheet_data['merged_cells']. Текущий размер: {len(sheet_data['merged_cells'])}")
            
            logger.info(f"Завершено извлечение объединенных ячеек с листа '{sheet_name}'. Найдено: {merged_count}")

            # Добавляем данные листа в результаты анализа
            analysis_results["sheets"].append(sheet_data)

        logger.info(f"Анализ Excel-файла '{file_path}' завершен.")
        return analysis_results

    except Exception as e:
        logger.error(f"Ошибка при анализе Excel-файла '{file_path}': {e}", exc_info=True)
        # Возвращаем пустой словарь или поднимаем исключение
        # В реальном приложении лучше поднимать пользовательское исключение
        raise # Повторно поднимаем исключение для обработки выше

# --- Функция для группировки адресов ячеек в диапазоны (заглушка) ---
# def _group_addresses(addresses: List[str]) -> str:
#     """
#     Группирует список адресов ячеек в строку диапазонов.
#     Например: ['A1', 'A2', 'A3', 'B1'] -> 'A1:A3 B1'
#     Это сложная задача, требующая алгоритмов.
#     """
#     # TODO: Реализовать алгоритм группировки
#     # Пока возвращаем объединение через пробел
#     return " ".join(addresses)

# Пример использования (если файл запускается напрямую)
if __name__ == "__main__":
    import sys
    if len(sys.argv) != 2:
        print("Использование: python logic_documentation.py <excel_file_path>")
        sys.exit(1)

    file_path = sys.argv[1]
    try:
        results = analyze_excel_file(file_path)
        print(f"Анализ завершен. Результаты для {results['project_name']}:")
        print(f"  - Листов: {len(results['sheets'])}")
        for sheet in results['sheets']:
            print(f"    - Лист '{sheet['name']}':")
            print(f"      - Ячеек с данными: {len(sheet['raw_data'])}")
            print(f"      - Формул: {len(sheet['formulas'])}")
            print(f"      - Стилей: {len(sheet['styles'])}")
            print(f"      - Диаграмм: {len(sheet['charts'])}")
            print(f"      - Объединенных ячеек: {len(sheet['merged_cells'])}")
    except Exception as e:
        print(f"Ошибка при анализе: {e}")
        sys.exit(1)
