# scripts/create_test_excel.py
"""
Скрипт для создания тестового Excel-файла для проверки анализатора.
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
import sys

def create_test_excel_file(output_path: str = None):
    """
    Создает тестовый Excel-файл с различными типами данных и формулами.
    
    Args:
        output_path (str): Путь для сохранения файла (опционально)
    """
    
    # Определяем путь относительно корня проекта
    if output_path is None:
        # Получаем путь к корню проекта (директория, где находится этот скрипт)
        script_dir = Path(__file__).parent
        project_root = script_dir.parent  # Поднимаемся на уровень выше (из scripts/ в корень)
        output_path = project_root / "data" / "samples" / "test_sample.xlsx"
    else:
        output_path = Path(output_path)
    
    # Создаем директорию, если её нет
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Создаем книгу Excel
    wb = Workbook()
    
    # Лист 1: Простые данные и формулы
    ws1 = wb.active
    ws1.title = "Sales_Data"
    
    # Заголовки
    headers = ["Product", "Q1_Sales", "Q2_Sales", "Q3_Sales", "Q4_Sales", "Total_Sales", "Average_Sales"]
    ws1.append(headers)
    
    # Данные
    data = [
        ["Product A", 100, 150, 200, 180, "=SUM(B2:E2)", "=AVERAGE(B2:E2)"],
        ["Product B", 80, 120, 90, 110, "=SUM(B3:E3)", "=AVERAGE(B3:E3)"],
        ["Product C", 200, 180, 190, 210, "=SUM(B4:E4)", "=AVERAGE(B4:E4)"],
        ["Product D", 50, 60, 70, 80, "=SUM(B5:E5)", "=AVERAGE(B5:E5)"],
    ]
    
    for row in data:
        ws1.append(row)
    
    # Лист 2: Смешанные типы данных
    ws2 = wb.create_sheet("Mixed_Data")
    
    # Заголовки
    ws2.append(["ID", "Name", "Date", "Active", "Score", "Notes"])
    
    # Данные с разными типами
    mixed_data = [
        [1, "Alice", datetime(2024, 1, 15), True, 95.5, "Top performer"],
        [2, "Bob", datetime(2024, 2, 20), False, 78.0, ""],
        [3, "Charlie", datetime(2024, 3, 10), True, 88.2, "Needs improvement"],
        [4, "Diana", datetime(2024, 4, 5), True, 92.7, "Consistent"],
        [5, "Eve", datetime(2024, 5, 12), False, 65.3, "New hire"],
    ]
    
    for row in mixed_data:
        ws2.append(row)
    
    # Лист 3: Формулы с ссылками между листами
    ws3 = wb.create_sheet("Summary")
    
    # Заголовки
    ws3.append(["Metric", "Value", "Source"])
    
    # Формулы, ссылающиеся на другие листы
    summary_data = [
        ["Total Q1 Sales", "=SUM(Sales_Data!B2:B5)", "Sales_Data"],
        ["Average Score", "=AVERAGE(Mixed_Data!E2:E6)", "Mixed_Data"],
        ["Active Employees", "=COUNTIF(Mixed_Data!D2:D6,TRUE)", "Mixed_Data"],
        ["Max Sales", "=MAX(Sales_Data!F2:F5)", "Sales_Data"],
    ]
    
    for row in summary_data:
        ws3.append(row)
    
    # Лист 4: Пустые значения и ошибки
    ws4 = wb.create_sheet("Edge_Cases")
    
    # Заголовки
    ws4.append(["Value1", "Value2", "Result", "Notes"])
    
    # Данные с пустыми значениями и делением на ноль
    edge_data = [
        [10, 2, "=A2/B2", "Normal division"],
        [20, 0, "=A3/B3", "Division by zero"],
        [30, "", "=A4+B4", "Empty cell reference"],
        ["", "", "=A5/B5", "Both empty"],
        [5, 3, "=A6*B6", "Multiplication"],
    ]
    
    for row in edge_data:
        ws4.append(row)
    
    # Сохраняем файл
    wb.save(output_path)
    print(f"Тестовый Excel-файл создан: {output_path.absolute()}")

if __name__ == "__main__":
    # Если скрипт запущен напрямую, создаем файл в стандартном месте
    create_test_excel_file()